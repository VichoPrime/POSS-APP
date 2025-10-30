from flask import Flask, request, jsonify, session, render_template, send_file
from models.article import Article, Category
from models.user import User
from sqlalchemy import text
from models.sale import Sale, SaleItem, Turno, SuspendedSale, Devolucion
from models.inventory_loss import InventoryLoss
from models.physical_inventory import PhysicalInventory
from models.discount import Discount, Promotion, SaleDiscount
from models.history import ProductHistory, PhysicalCountHistory
from models import db
from flask_cors import CORS
from functools import wraps
import os
import mimetypes
from datetime import datetime, timezone, timedelta
import uuid
import pytz
from sqlalchemy import create_engine, func, case
from sqlalchemy.orm import sessionmaker
import json
from werkzeug.utils import secure_filename
import openpyxl
import time
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import io

# Configuración para upload de archivos
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'uploads')
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# Configurar Flask con carpeta static
app = Flask(__name__, static_folder='static', static_url_path='/static')

# CONFIGURACIÓN SIMPLIFICADA PARA DESARROLLO LOCAL
app.config.update(
    SECRET_KEY='brian123',
    SESSION_COOKIE_SAMESITE='Lax',    # Cambiar a Lax para localhost
    SESSION_COOKIE_SECURE=False,       # False para HTTP
    SESSION_COOKIE_HTTPONLY=False,     # False para debugging
    SESSION_COOKIE_DOMAIN=None,        # Sin restricción de dominio
    SESSION_COOKIE_NAME='session',
    SESSION_COOKIE_PATH='/',
    UPLOAD_FOLDER=UPLOAD_FOLDER,
    MAX_CONTENT_LENGTH=16 * 1024 * 1024  # 16MB máx
)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB máx

# CORS simplificado para desarrollo local
CORS(app, 
     supports_credentials=True,
     origins=['http://localhost:3000', 'http://127.0.0.1:3000', 'http://localhost:3001', 'http://127.0.0.1:3001'],
     allow_headers=['Content-Type', 'Authorization'],
     methods=['GET', 'POST', 'PUT', 'DELETE', 'OPTIONS'])

# Configuración para upload de archivos
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'uploads')
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max

# Crear directorio de uploads si no existe
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

import os
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.abspath(os.path.join(os.path.dirname(__file__), 'instance', 'database.db'))
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

# Asegurar que el import incluya Devolucion
with app.app_context():
    try:
        db.create_all()  # Esto creará solo las tablas/columnas faltantes
        print("Base de datos actualizada")
    except Exception as e:
        print(f"Error actualizando BD: {e}")

# Definir el decorador login_required
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        print(f"🔒 login_required check for {request.endpoint}")
        print(f"🔒 Session data: {dict(session)}")
        print(f"🔒 Cookies: {dict(request.cookies)}")
        
        if 'user_id' not in session:
            print(f"❌ No user_id in session, returning 401")
            return jsonify({'error': 'Autenticación requerida'}), 401
        
        print(f"✅ User authenticated: {session['user_id']}")
        return f(*args, **kwargs)
    return decorated_function

# Decorador para verificar permisos específicos
def permission_required(permission):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                return jsonify({'error': 'Usuario no autenticado'}), 401
            
            user = User.query.get(session['user_id'])
            if not user:
                return jsonify({'error': 'Usuario no válido'}), 401
            
            if not user.has_permission(permission):
                return jsonify({'error': 'No tienes permisos para esta acción'}), 403
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# Decorador para administradores únicamente
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Usuario no autenticado'}), 401
        
        user = User.query.get(session['user_id'])
        if not user or not user.is_admin:
            return jsonify({'error': 'Acceso restringido a administradores'}), 403
        
        return f(*args, **kwargs)
    return decorated_function

# Nueva función para obtener o crear turno activo
def get_or_create_active_turno(user_id):
    turno_activo = Turno.query.filter_by(user_id=user_id, activo=True).first()
    if not turno_activo:
        # Crear nuevo turno con la zona horaria de Chile
        chile_tz = pytz.timezone('America/Santiago')
        turno_activo = Turno(
            user_id=user_id,
            fecha_inicio=datetime.now(chile_tz)
        )
        db.session.add(turno_activo)
        db.session.flush()
    return turno_activo

# Funciones para historial de auditoría
def log_product_change(article_id, user_id, action, description, old_values=None, new_values=None):
    """Registra un cambio en el historial de productos"""
    try:
        history_entry = ProductHistory(
            article_id=article_id,
            user_id=user_id,
            action=action,
            description=description,
            old_values=json.dumps(old_values) if old_values else None,
            new_values=json.dumps(new_values) if new_values else None
        )
        db.session.add(history_entry)
        db.session.flush()
        print(f"📝 Historial registrado: {action} en producto {article_id} por usuario {user_id}")
    except Exception as e:
        print(f"❌ Error registrando historial: {e}")

def log_physical_count(article_id, user_id, old_stock, new_stock, observation=None):
    """Registra un conteo físico en el historial"""
    try:
        difference = new_stock - old_stock
        count_entry = PhysicalCountHistory(
            article_id=article_id,
            user_id=user_id,
            old_stock=old_stock,
            new_stock=new_stock,
            difference=difference,
            observation=observation
        )
        db.session.add(count_entry)
        db.session.flush()
        print(f"📊 Conteo físico registrado: producto {article_id}, {old_stock} -> {new_stock}")
    except Exception as e:
        print(f"❌ Error registrando conteo físico: {e}")

# Función para validar archivos de imagen
def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# =====================
# RUTAS DE PÁGINAS
# =====================

@app.route('/')
def home(): 
    return render_template('caja.html')

@app.route('/buscar-productos')
def buscar_productos():
    return render_template('articles.html')

@app.route('/historial-ventas')
def historial_ventas():
    return render_template('ventas.html')

@app.route('/turnos')
@login_required
def turnos():
    return render_template('turnos.html')

# =====================
# AUTENTICACIÓN
# =====================

@app.route('/register', methods=['POST'])
def register_user():
    data = request.get_json()
    
    if not data:
        return jsonify({'error': 'No se proporcionaron datos'}), 400
    
    required_fields = ['username', 'email', 'password']
    if not all(field in data for field in required_fields):
        return jsonify({'error': 'Faltan campos requeridos'}), 400
    
    if User.query.filter_by(email=data['email']).first() is not None:
        return jsonify({'error': 'El email ya esta registrado'}), 400
    
    new_user = User(username=data['username'], email=data['email'])
    new_user.set_password(data['password'])
    db.session.add(new_user)
    db.session.commit()
    return jsonify({
        'message': f'usuario {new_user.username} registrado con exito'
    }), 201

@app.route('/login', methods=['POST'])
def login_user():
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'No se proporcionaron datos'}), 400
        
        required_fields = ['email', 'password']
        if not all(field in data for field in required_fields):
            return jsonify({'error': 'Faltan campos requeridos: email y password'}), 400
        
        # Buscar usuario
        user = User.query.filter_by(email=data['email']).first()
        if user is None:
            return jsonify({'error': 'Usuario no encontrado'}), 401
        
        # Verificar contraseña
        if not user.check_password(data['password']):
            return jsonify({'error': 'Contraseña incorrecta'}), 401

        # Crear sesión
        session.clear()  # Limpiar sesión anterior
        session['user_id'] = user.id
        session.permanent = True  # Hacer la sesión permanente
        session.modified = True  # Forzar guardado de sesión
        
        # Debug: Imprimir información de la sesión
        print(f"🔐 Login exitoso - User ID: {user.id}")
        print(f"🔐 Session ID: {session.get('_id', 'No ID')}")
        print(f"🔐 Session data: {dict(session)}")
        print(f"🔐 Session permanent: {session.permanent}")
        
        # Crear o obtener turno activo
        turno_activo = get_or_create_active_turno(user.id)
        db.session.commit()
        
        return jsonify({
            'message': f'Bienvenido {user.username}',
            'user_id': user.id,
            'username': user.username,
            'email': user.email,
            'turno_id': turno_activo.id,
            'turno_inicio': turno_activo.fecha_inicio.strftime('%Y-%m-%d %H:%M:%S')
        })
        
    except Exception as e:
        print(f"Error en login: {str(e)}")
        return jsonify({'error': f'Error interno del servidor: {str(e)}'}), 500

@app.route('/check-auth', methods=['GET'])
def check_auth():
    try:
        print(f"🔍 Check-auth - Session data: {dict(session)}")
        print(f"🔍 Request headers: {dict(request.headers)}")
        print(f"🔍 Request cookies: {request.cookies}")
        
        if 'user_id' in session:
            user = User.query.get(session['user_id'])
            if user:
                return jsonify({
                    'logged_in': True,
                    'user_id': user.id,
                    'username': user.username,
                    'email': user.email
                })
        
        return jsonify({'logged_in': False})
        
    except Exception as e:
        print(f"Error en check_auth: {str(e)}")
        return jsonify({'logged_in': False})

# Endpoint de debug para verificar sesiones
@app.route('/debug-session', methods=['GET'])
def debug_session():
    return jsonify({
        'session_data': dict(session),
        'has_user_id': 'user_id' in session,
        'user_id': session.get('user_id'),
        'session_id': session.get('_id'),
        'cookies_received': dict(request.cookies),
        'headers': dict(request.headers)
    })

# Endpoint para test simple de sesión
@app.route('/test-session', methods=['POST'])
def test_session():
    """Endpoint simple para probar si las sesiones funcionan"""
    session['test_value'] = 'session_works'
    session.modified = True
    return jsonify({
        'message': 'Session test value set',
        'session_data': dict(session),
        'test_value': session.get('test_value')
    })

@app.route('/test-session', methods=['GET'])
def get_test_session():
    """Verificar si el valor de test persiste"""
    return jsonify({
        'message': 'Reading session test',
        'session_data': dict(session),
        'test_value': session.get('test_value'),
        'has_test': 'test_value' in session
    })

@app.route('/logout', methods=['POST'])
def logout_user():
    try:
        # Verificar si hay turno activo y cerrarlo
        if 'user_id' in session:
            turno_activo = Turno.query.filter_by(user_id=session['user_id'], activo=True).first()
            if turno_activo:
                turno_activo.fecha_cierre = datetime.utcnow()
                turno_activo.activo = False
                db.session.commit()
        
        session.pop('user_id', None)
        return jsonify({'message': 'Sesión cerrada con éxito'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': 'Sesión cerrada con éxito'})

# =====================
# UPLOAD DE IMÁGENES
# =====================

@app.route('/upload-image', methods=['POST'])
@login_required
def upload_image():
    try:
        if 'image' not in request.files:
            return jsonify({'error': 'No se proporcionó archivo'}), 400
        
        file = request.files['image']
        if file.filename == '':
            return jsonify({'error': 'No se seleccionó archivo'}), 400
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            # Agregar timestamp para evitar conflictos
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
            filename = timestamp + filename
            
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            
            # URL absoluta para acceder a la imagen desde el frontend
            image_url = f'http://localhost:5000/static/uploads/{filename}'
            
            return jsonify({
                'success': True,
                'image_url': image_url,
                'filename': filename
            })
        else:
            return jsonify({'error': 'Tipo de archivo no permitido'}), 400
            
    except Exception as e:
        print(f"Error uploading image: {str(e)}")
        return jsonify({'error': 'Error al subir imagen'}), 500

# Servir archivos de imágenes con MIME types correctos
@app.route('/static/uploads/<filename>')
def uploaded_file(filename):
    from flask import send_from_directory, make_response
    
    try:
        print(f"🖼️ Serving image: {filename}")
        
        # Obtener el MIME type correcto basado en la extensión
        mime_type, _ = mimetypes.guess_type(filename)
        if mime_type is None:
            # Fallback para extensiones no reconocidas
            if filename.lower().endswith('.webp'):
                mime_type = 'image/webp'
            elif filename.lower().endswith(('.jpg', '.jpeg')):
                mime_type = 'image/jpeg'
            elif filename.lower().endswith('.png'):
                mime_type = 'image/png'
            elif filename.lower().endswith('.gif'):
                mime_type = 'image/gif'
            else:
                mime_type = 'application/octet-stream'
        
        print(f"📄 MIME type: {mime_type}")
        
        # Crear respuesta con MIME type correcto
        response = make_response(send_from_directory(app.config['UPLOAD_FOLDER'], filename))
        response.headers['Content-Type'] = mime_type
        response.headers['Cache-Control'] = 'public, max-age=3600'  # Cache por 1 hora
        
        # Headers CORS específicos para imágenes
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'GET'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        
        print(f"✅ Image served successfully: {filename}")
        return response
        
    except FileNotFoundError:
        print(f"❌ File not found: {filename}")
        return jsonify({'error': 'Archivo no encontrado'}), 404
    except Exception as e:
        print(f"❌ Error serving file {filename}: {str(e)}")
        return jsonify({'error': 'Error al servir archivo'}), 500

# =====================
# PRODUCTOS Y CATEGORÍAS
# =====================

@app.route('/articles', methods=['GET'])
def get_articles():
    try:
        category_id = request.args.get('category_id')
        
        if category_id:
            articles = Article.query.filter_by(category_id=category_id, activo=True).all()
        else:
            articles = Article.query.filter_by(activo=True).all()
        
        articles_data = []
        for article in articles:
            # Obtener nombre de categoría
            category_name = "Sin categoría"
            if article.category_id:
                category = Category.query.get(article.category_id)
                if category:
                    category_name = category.name
            
            article_data = {
                'id': article.id,
                'title': article.title,
                'content': article.content,
                'precio': float(article.precio),
                'stock': article.stock,
                'stock_minimo': getattr(article, 'stock_minimo', 5),
                'is_low_stock': article.stock <= getattr(article, 'stock_minimo', 5),
                'rotacion_minima': article.rotacion_minima,
                'is_low_rotation': article.is_low_rotation() if article.rotacion_minima is not None else False,
                'barcode': article.codigo_barra,
                'category_id': article.category_id,
                'category_name': category_name,
                'image_url': article.image_url,
                'activo': article.activo,
                'unit_type': getattr(article, 'unit_type', 'unidades'),
                'peso_unitario': getattr(article, 'peso_unitario', None),
                'margen_ganancia': getattr(article, 'margen_ganancia', 0),
                'precio_costo': article.get_precio_costo() if hasattr(article, 'get_precio_costo') else article.precio
            }
            articles_data.append(article_data)
        
        return jsonify(articles_data)
        
    except Exception as e:
        print(f"Error al obtener artículos: {str(e)}")
        return jsonify({'error': 'Error al obtener artículos'}), 500

@app.route('/articles/low-rotation', methods=['GET'])
@login_required
def get_low_rotation_articles():
    """Obtiene productos con baja rotación"""
    try:
        # Obtener todos los productos con rotación mínima configurada
        articles = Article.query.filter(
            Article.rotacion_minima.isnot(None),
            Article.activo == True
        ).all()
        
        # Filtrar artículos con baja rotación
        low_rotation_articles = [article for article in articles if article.is_low_rotation()]
        
        articles_data = []
        for article in low_rotation_articles:
            # Calcular ventas de los últimos 30 días
            thirty_days_ago = datetime.utcnow() - timedelta(days=30)
            total_sales = sum(
                item.quantity
                for item in article.article_sale_items
                if item.sale.fecha_venta >= thirty_days_ago
            )
            
            category_name = "Sin categoría"
            if article.category_id:
                category = Category.query.get(article.category_id)
                if category:
                    category_name = category.name
                    
            article_data = {
                'id': article.id,
                'title': article.title,
                'content': article.content,
                'precio': float(article.precio),
                'stock': article.stock,
                'rotacion_minima': article.rotacion_minima,
                'ventas_30_dias': total_sales,
                'category_name': category_name,
                'image_url': article.image_url
            }
            articles_data.append(article_data)
        
        return jsonify(articles_data)
        
    except Exception as e:
        print(f"Error al obtener productos con baja rotación: {str(e)}")
        return jsonify({'error': 'Error al obtener productos con baja rotación'}), 500

@app.route('/articles/stock-report', methods=['GET'])
@login_required
def get_stock_report():
    """Obtiene un reporte detallado de stock incluyendo pérdidas"""
    try:
        # Obtener productos con stock bajo
        low_stock_articles = Article.query.filter(
            Article.stock <= Article.stock_minimo,
            Article.activo == True
        ).all()
        
        stock_data = []
        for article in low_stock_articles:
            # Calcular pérdidas totales para este artículo
            total_losses = db.session.query(func.sum(InventoryLoss.cantidad_perdida)).filter(
                InventoryLoss.article_id == article.id
            ).scalar() or 0
            
            # Calcular pérdidas por tipo
            losses_by_type = db.session.query(
                InventoryLoss.tipo_perdida,
                func.sum(InventoryLoss.cantidad_perdida)
            ).filter(
                InventoryLoss.article_id == article.id
            ).group_by(InventoryLoss.tipo_perdida).all()
            
            losses_detail = {loss_type: amount for loss_type, amount in losses_by_type}
            
            category_name = article.category.name if article.category else 'Sin categoría'
            
            stock_data.append({
                'id': article.id,
                'title': article.title,
                'stock_actual': article.stock,
                'stock_minimo': article.stock_minimo,
                'unit_type': getattr(article, 'unit_type', 'unidades'),
                'category_name': category_name,
                'total_losses': total_losses,
                'losses_detail': {
                    'vencido': losses_detail.get('vencido', 0),
                    'dañado': losses_detail.get('dañado', 0)
                },
                'stock_status': 'critico' if article.stock == 0 else 'bajo'
            })
        
        # Estadísticas globales
        total_articles = Article.query.filter(Article.activo == True).count()
        low_stock_count = len(stock_data)
        out_of_stock = Article.query.filter(Article.stock == 0, Article.activo == True).count()
        
        # Total de pérdidas globales
        global_losses = db.session.query(
            InventoryLoss.tipo_perdida,
            func.sum(InventoryLoss.cantidad_perdida)
        ).group_by(InventoryLoss.tipo_perdida).all()
        
        global_losses_dict = {loss_type: amount for loss_type, amount in global_losses}
        
        return jsonify({
            'low_stock_articles': stock_data,
            'summary': {
                'total_articles': total_articles,
                'low_stock_count': low_stock_count,
                'out_of_stock': out_of_stock,
                'global_losses': {
                    'vencido': global_losses_dict.get('vencido', 0),
                    'dañado': global_losses_dict.get('dañado', 0),
                    'total': sum(global_losses_dict.values())
                }
            }
        })
        
    except Exception as e:
        print(f"Error al obtener reporte de stock: {str(e)}")
        return jsonify({'error': 'Error al obtener reporte de stock'}), 500

@app.route('/articles', methods=['POST'])
@permission_required('can_manage_products')
def create_article():
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'No se proporcionaron datos'}), 400
        
        # Los campos requeridos ahora son menos estrictos
        required_fields = ['title', 'precio']
        missing_fields = [field for field in required_fields if field not in data]
        
        if missing_fields:
            return jsonify({'error': f'Faltan campos requeridos: {", ".join(missing_fields)}'}), 400
        
        # Validar que la categoría exista si se proporciona
        category_id = data.get('category_id')
        if category_id:
            category = Category.query.get(category_id)
            if not category:
                return jsonify({'error': f'Categoría con ID {category_id} no encontrada'}), 400
        
        # Validar tipos de datos
        try:
            stock = float(data.get('stock', 0))  # Cambiar a float para soportar peso
            # Convertir precios a enteros para pesos chilenos (sin decimales)
            precio = int(float(data['precio']))
            stock_minimo = int(data.get('stock_minimo', 5))
            rotacion_minima = int(data['rotacion_minima']) if data.get('rotacion_minima') else None
            unit_type = data.get('unit_type', 'unidades')
            peso_unitario = data.get('peso_unitario')
            margen_ganancia = data.get('margen_ganancia', 0)
            
            if peso_unitario is not None:
                peso_unitario = float(peso_unitario)
            
            if margen_ganancia is not None:
                margen_ganancia = int(float(margen_ganancia))
                
        except (ValueError, TypeError):
            return jsonify({'error': 'Tipos de datos inválidos'}), 400
        
        # Crear el artículo
        new_article = Article(
            title=data['title'],
            content=data.get('content', ''),
            image_url=data.get('image_url', ''),
            stock=stock,
            stock_minimo=stock_minimo,
            codigo_barra=data.get('codigo_barra', ''),
            precio=precio,
            activo=data.get('activo', True),
            category_id=category_id,
            unit_type=unit_type,
            peso_unitario=peso_unitario,
            margen_ganancia=margen_ganancia,
            rotacion_minima=rotacion_minima
        )

        db.session.add(new_article)
        db.session.commit()
        
        # Registrar en historial
        product_data = {
            'title': new_article.title,
            'precio': new_article.precio,
            'stock': new_article.stock,
            'category_id': new_article.category_id,
            'unit_type': new_article.unit_type
        }
        log_product_change(
            article_id=new_article.id,
            user_id=session['user_id'],
            action='created',
            description=f'Producto creado: {new_article.title}',
            new_values=product_data
        )
        db.session.commit()

        return jsonify({
            'id': new_article.id,
            'title': new_article.title,
            'content': new_article.content,
            'image_url': new_article.image_url,
            'stock': new_article.stock,
            'stock_minimo': new_article.stock_minimo,
            'is_low_stock': new_article.stock <= new_article.stock_minimo,
            'codigo_barra': new_article.codigo_barra,
            'precio': new_article.precio,
            'activo': new_article.activo,
            'unit_type': new_article.unit_type,
            'peso_unitario': new_article.peso_unitario,
            'margen_ganancia': new_article.margen_ganancia,
            'precio_costo': new_article.get_precio_costo(),
            'codigo_barra': new_article.codigo_barra,
            'precio': new_article.precio,
            'activo': new_article.activo,
            'unit_type': new_article.unit_type,
            'peso_unitario': new_article.peso_unitario,
            'category': category.name if category_id and category else 'Sin categoría'
        }), 201
    except Exception as e:
        db.session.rollback()
        print(f"Error creating article: {str(e)}")
        return jsonify({'error': 'Error al crear el artículo'}), 500

@app.route('/articles/<int:id>', methods=['PUT'])
@permission_required('can_manage_products')
def update_article(id):
    article = Article.query.get_or_404(id)
    data = request.get_json()
    
    if not data:
        return jsonify({'error': 'No se proporcionaron datos'}), 400
    
    # Capturar valores anteriores para historial
    old_values = {
        'title': article.title,
        'content': article.content,
        'stock': article.stock,
        'precio': article.precio,
        'activo': article.activo,
        'unit_type': getattr(article, 'unit_type', 'unidades'),
        'peso_unitario': getattr(article, 'peso_unitario', None),
        'category_id': article.category_id
    }
    
    changes_made = []
    
    # Actualizar campos si están presentes en los datos
    if 'title' in data and data['title'] != article.title:
        changes_made.append(f"Título: '{article.title}' → '{data['title']}'")
        article.title = data['title']
    if 'content' in data and data['content'] != article.content:
        changes_made.append(f"Descripción modificada")
        article.content = data['content']
    if 'image_url' in data:
        article.image_url = data['image_url']
    if 'stock' in data and int(float(data['stock'])) != article.stock:
        changes_made.append(f"Stock: {article.stock} → {int(float(data['stock']))}")
        article.stock = int(float(data['stock']))
    if 'codigo_barra' in data:
        article.codigo_barra = data['codigo_barra']
    if 'precio' in data and int(float(data['precio'])) != article.precio:
        changes_made.append(f"Precio: ${article.precio} → ${int(float(data['precio']))}")
        article.precio = int(float(data['precio']))  # Convertir a entero para pesos chilenos
    if 'margen_ganancia' in data and int(float(data['margen_ganancia'])) != article.margen_ganancia:
        changes_made.append(f"Margen: ${article.margen_ganancia} → ${int(float(data['margen_ganancia']))}")
        article.margen_ganancia = int(float(data['margen_ganancia']))  # Convertir a entero
    if 'activo' in data and data['activo'] != article.activo:
        status = "activado" if data['activo'] else "desactivado"
        changes_made.append(f"Producto {status}")
        article.activo = data['activo']
    if 'unit_type' in data:
        article.unit_type = data['unit_type']
    if 'peso_unitario' in data:
        article.peso_unitario = float(data['peso_unitario']) if data['peso_unitario'] else None
    if 'stock_minimo' in data:
        article.stock_minimo = int(data['stock_minimo']) if data['stock_minimo'] else 5
    if 'rotacion_minima' in data:
        article.rotacion_minima = int(data['rotacion_minima']) if data['rotacion_minima'] else None
    if 'category_id' in data:
        if data['category_id']:
            category = Category.query.get(data['category_id'])
            if not category:
                return jsonify({'error': 'Categoría no encontrada'}), 400
            article.category_id = data['category_id']
        else:
            article.category_id = None
    
    # Capturar valores nuevos para historial
    new_values = {
        'title': article.title,
        'content': article.content,
        'stock': article.stock,
        'precio': article.precio,
        'activo': article.activo,
        'unit_type': getattr(article, 'unit_type', 'unidades'),
        'peso_unitario': getattr(article, 'peso_unitario', None),
        'category_id': article.category_id
    }
    
    db.session.commit()
    
    # Registrar en historial solo si hubo cambios
    if changes_made:
        description = f"Producto actualizado: {'; '.join(changes_made)}"
        log_product_change(
            article_id=article.id,
            user_id=session['user_id'],
            action='updated',
            description=description,
            old_values=old_values,
            new_values=new_values
        )
        db.session.commit()

    return jsonify({
        'id': article.id,
        'title': article.title,
        'content': article.content,
        'image_url': article.image_url,
        'stock': article.stock,
        'codigo_barra': article.codigo_barra,
        'precio': article.precio,
        'activo': article.activo,
        'unit_type': getattr(article, 'unit_type', 'unidades'),
        'peso_unitario': getattr(article, 'peso_unitario', None),
        'category': article.category.name if article.category else None
    })

@app.route('/articles/<int:id>', methods=['DELETE'])
@permission_required('can_manage_products')
def delete_article(id):
    article = Article.query.get_or_404(id)
    db.session.delete(article)
    db.session.commit()
    return jsonify({
        'message': f'Articulo eliminado con exito, se elimino {article.title}'
    }), 200

@app.route('/article/<int:article_id>', methods=['GET'])
def view_article(article_id):
    article = Article.query.get_or_404(article_id)
    return jsonify({
        'id': article.id,
        'title': article.title,
        'content': article.content
    })

@app.route('/articles/barcode/<string:barcode>', methods=['GET'])
def get_article_by_barcode(barcode):
    try:
        article = Article.query.filter_by(codigo_barra=barcode, activo=True).first()
        
        if not article:
            return jsonify({'error': 'Artículo no encontrado'}), 404
        
        # Obtener nombre de categoría
        category_name = "Sin categoría"
        if article.category_id:
            category = Category.query.get(article.category_id)
            if category:
                category_name = category.name
        
        article_data = {
            'id': article.id,
            'title': article.title,
            'content': article.content,
            'precio': float(article.precio),
            'stock': article.stock,
            'barcode': article.codigo_barra,
            'category_id': article.category_id,
            'category_name': category_name,
            'image_url': article.image_url,
            'activo': article.activo
        }
        
        return jsonify(article_data)
        
    except Exception as e:
        print(f"Error al buscar artículo: {str(e)}")
        return jsonify({'error': 'Error al buscar artículo'}), 500

@app.route('/categories', methods=['GET'])
def get_categories():
    categories = Category.query.all()
    return jsonify([{
        'id': cat.id,
        'name': cat.name
    } for cat in categories])

@app.route('/categories', methods=['POST'])
def create_category():
    data = request.get_json()
    name = data['name'].strip().capitalize()

    if Category.query.filter_by(name=name).first():
        return jsonify({'error': 'La categoría ya existe'}), 400

    category = Category(name=name)
    db.session.add(category)
    db.session.commit()

    return jsonify({'id': category.id, 'name': category.name}), 201

def update_frequent_products_category():
    """Actualiza la categoría 'Productos Frecuentes' con los 3 productos más vendidos"""
    try:
        # Buscar o crear la categoría "Productos Frecuentes"
        frequent_category = Category.query.filter_by(name='Productos Frecuentes').first()
        if not frequent_category:
            frequent_category = Category(name='Productos Frecuentes')
            db.session.add(frequent_category)
            db.session.commit()
        
        # Primero, quitar todos los productos de esta categoría
        Article.query.filter_by(category_id=frequent_category.id).update({'category_id': None})
        
        # Obtener los 3 productos más vendidos
        most_sold = db.session.query(
            SaleItem.article_id,
            db.func.sum(SaleItem.quantity).label('total_sold')
        ).group_by(SaleItem.article_id).order_by(
            db.func.sum(SaleItem.quantity).desc()
        ).limit(3).all()
        
        # Asignar los productos más vendidos a la categoría frecuentes
        for item in most_sold:
            article = Article.query.get(item.article_id)
            if article:
                article.category_id = frequent_category.id
        
        db.session.commit()
        return frequent_category.id
    except Exception as e:
        print(f"Error updating frequent products category: {e}")
        db.session.rollback()
        return None

@app.route('/categories/update-frequent', methods=['POST'])
def update_frequent_category():
    """Endpoint para actualizar manualmente la categoría de productos frecuentes"""
    category_id = update_frequent_products_category()
    if category_id:
        return jsonify({'message': 'Categoría de productos frecuentes actualizada', 'category_id': category_id}), 200
    else:
        return jsonify({'error': 'Error al actualizar la categoría'}), 500

# =====================
# PÉRDIDAS DE INVENTARIO
# =====================

@app.route('/inventory-losses', methods=['POST'])
@permission_required('can_manage_inventory_losses')
def create_inventory_loss():
    try:
        data = request.get_json()
        
        # Validar datos requeridos
        required_fields = ['article_id', 'cantidad_perdida', 'tipo_perdida']
        for field in required_fields:
            if field not in data or data[field] is None:
                return jsonify({'error': f'Campo requerido: {field}'}), 400
        
        # Validar que el artículo existe
        article = Article.query.get(data['article_id'])
        if not article:
            return jsonify({'error': 'Artículo no encontrado'}), 404
            
        # Validar cantidad
        cantidad_perdida = float(data['cantidad_perdida'])
        if cantidad_perdida <= 0:
            return jsonify({'error': 'La cantidad debe ser mayor a 0'}), 400
            
        # Validar que hay suficiente stock
        if cantidad_perdida > article.stock:
            return jsonify({'error': f'Stock insuficiente. Stock actual: {article.stock}'}), 400
            
        # Validar tipo de pérdida
        if data['tipo_perdida'] not in ['vencido', 'dañado']:
            return jsonify({'error': 'Tipo de pérdida debe ser "vencido" o "dañado"'}), 400
        
        # Crear registro de pérdida
        loss = InventoryLoss(
            article_id=data['article_id'],
            cantidad_perdida=cantidad_perdida,
            tipo_perdida=data['tipo_perdida'],
            motivo=data.get('motivo', ''),
            usuario_registro=session.get('username', 'Desconocido')
        )
        
        # Descontar del stock del artículo
        article.stock -= cantidad_perdida
        
        db.session.add(loss)
        db.session.commit()
        
        return jsonify({
            'message': 'Pérdida registrada exitosamente',
            'loss': loss.to_dict(),
            'new_stock': article.stock
        }), 201
        
    except ValueError:
        return jsonify({'error': 'Cantidad debe ser un número válido'}), 400
    except Exception as e:
        db.session.rollback()
        print(f"Error al registrar pérdida: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500

@app.route('/inventory-losses', methods=['GET'])
@permission_required('can_manage_inventory_losses')
def get_inventory_losses():
    try:
        # Parámetros de filtro opcionales
        article_id = request.args.get('article_id', type=int)
        tipo_perdida = request.args.get('tipo_perdida')
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        
        # Construir query base
        query = InventoryLoss.query
        
        # Aplicar filtros
        if article_id:
            query = query.filter(InventoryLoss.article_id == article_id)
            
        if tipo_perdida:
            query = query.filter(InventoryLoss.tipo_perdida == tipo_perdida)
            
        if fecha_desde:
            fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
            query = query.filter(InventoryLoss.fecha_registro >= fecha_desde_dt)
            
        if fecha_hasta:
            fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d')
            query = query.filter(InventoryLoss.fecha_registro <= fecha_hasta_dt)
        
        # Ordenar por fecha más reciente
        query = query.order_by(InventoryLoss.fecha_registro.desc())
        
        # Paginar
        losses_paginated = query.paginate(
            page=page, 
            per_page=per_page, 
            error_out=False
        )
        
        # Obtener totales por tipo para estadísticas
        totals_query = InventoryLoss.query
        if article_id:
            totals_query = totals_query.filter(InventoryLoss.article_id == article_id)
            
        total_vencido = totals_query.filter(InventoryLoss.tipo_perdida == 'vencido').with_entities(func.sum(InventoryLoss.cantidad_perdida)).scalar() or 0
        total_dañado = totals_query.filter(InventoryLoss.tipo_perdida == 'dañado').with_entities(func.sum(InventoryLoss.cantidad_perdida)).scalar() or 0
        
        return jsonify({
            'losses': [loss.to_dict() for loss in losses_paginated.items],
            'pagination': {
                'page': page,
                'pages': losses_paginated.pages,
                'per_page': per_page,
                'total': losses_paginated.total
            },
            'totals': {
                'vencido': total_vencido,
                'dañado': total_dañado,
                'total': total_vencido + total_dañado
            }
        })
        
    except Exception as e:
        print(f"Error al obtener pérdidas: {str(e)}")
        return jsonify({'error': 'Error al obtener pérdidas'}), 500

@app.route('/inventory-losses/<int:loss_id>', methods=['DELETE'])
@permission_required('can_manage_inventory_losses')
def delete_inventory_loss(loss_id):
    try:
        loss = InventoryLoss.query.get_or_404(loss_id)
        article = loss.article
        
        # Restaurar el stock
        article.stock += loss.cantidad_perdida
        
        db.session.delete(loss)
        db.session.commit()
        
        return jsonify({
            'message': 'Pérdida eliminada y stock restaurado',
            'restored_stock': article.stock
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error al eliminar pérdida: {str(e)}")
        return jsonify({'error': 'Error al eliminar pérdida'}), 500

# =====================
# CONTEO FÍSICO DE INVENTARIO
# =====================

@app.route('/physical-inventory/start', methods=['POST'])
@login_required
def start_physical_inventory():
    """Inicia una nueva sesión de conteo físico"""
    try:
        import uuid
        
        # Generar ID único para la sesión de conteo
        session_id = str(uuid.uuid4())
        
        # Obtener todos los productos activos
        products = Article.query.filter(Article.activo == True).all()
        
        if not products:
            return jsonify({'error': 'No hay productos activos para contar'}), 400
        
        # Crear registros de conteo para todos los productos
        conteos = []
        for product in products:
            conteo = PhysicalInventory(
                conteo_session_id=session_id,
                article_id=product.id,
                cantidad_sistema=product.stock,
                usuario_conteo=session.get('username', 'Desconocido'),
                estado='pendiente'
            )
            conteos.append(conteo)
            db.session.add(conteo)
        
        db.session.commit()
        
        return jsonify({
            'message': 'Sesión de conteo físico iniciada',
            'session_id': session_id,
            'total_products': len(conteos)
        }), 201
        
    except Exception as e:
        db.session.rollback()
        print(f"Error al iniciar conteo físico: {str(e)}")
        return jsonify({'error': 'Error al iniciar conteo físico'}), 500

@app.route('/physical-inventory/<session_id>', methods=['GET'])
@login_required
def get_physical_inventory_session(session_id):
    """Obtiene todos los productos de una sesión de conteo"""
    try:
        conteos = PhysicalInventory.query.filter(
            PhysicalInventory.conteo_session_id == session_id
        ).order_by(PhysicalInventory.article_id).all()
        
        if not conteos:
            return jsonify({'error': 'Sesión de conteo no encontrada'}), 404
        
        # Calcular estadísticas
        total_productos = len(conteos)
        contados = len([c for c in conteos if c.estado == 'contado'])
        con_diferencias = len([c for c in conteos if c.diferencia is not None and abs(c.diferencia) > 0.01])
        
        return jsonify({
            'conteos': [conteo.to_dict() for conteo in conteos],
            'estadisticas': {
                'total_productos': total_productos,
                'contados': contados,
                'pendientes': total_productos - contados,
                'con_diferencias': con_diferencias,
                'progreso': round((contados / total_productos) * 100, 2) if total_productos > 0 else 0
            }
        })
        
    except Exception as e:
        print(f"Error al obtener sesión de conteo: {str(e)}")
        return jsonify({'error': 'Error al obtener sesión de conteo'}), 500

@app.route('/physical-inventory/<int:conteo_id>', methods=['PUT'])
@login_required
def update_physical_count(conteo_id):
    """Actualiza la cantidad física contada de un producto"""
    try:
        data = request.get_json()
        cantidad_fisica = data.get('cantidad_fisica')
        notas = data.get('notas', '')
        
        if cantidad_fisica is None or cantidad_fisica < 0:
            return jsonify({'error': 'Cantidad física debe ser mayor o igual a 0'}), 400
        
        conteo = PhysicalInventory.query.get_or_404(conteo_id)
        
        # Actualizar los datos del conteo
        conteo.cantidad_fisica = float(cantidad_fisica)
        conteo.calculate_difference()  # Calcula automáticamente la diferencia
        conteo.estado = 'contado'
        conteo.notas = notas
        
        db.session.commit()
        
        return jsonify({
            'message': 'Conteo actualizado exitosamente',
            'conteo': conteo.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error al actualizar conteo: {str(e)}")
        return jsonify({'error': 'Error al actualizar conteo'}), 500

@app.route('/physical-inventory/<session_id>/adjust', methods=['POST'])
@login_required
def apply_inventory_adjustments(session_id):
    """Aplica los ajustes de inventario basados en el conteo físico"""
    try:
        data = request.get_json()
        adjust_all = data.get('adjust_all', False)  # Si True, ajusta todos; si False, solo los seleccionados
        selected_ids = data.get('selected_ids', [])  # IDs específicos a ajustar
        
        # Obtener conteos de la sesión
        query = PhysicalInventory.query.filter(
            PhysicalInventory.conteo_session_id == session_id,
            PhysicalInventory.estado == 'contado',
            PhysicalInventory.diferencia != None
        )
        
        if not adjust_all and selected_ids:
            query = query.filter(PhysicalInventory.id.in_(selected_ids))
        
        conteos_a_ajustar = query.all()
        
        if not conteos_a_ajustar:
            return jsonify({'error': 'No hay conteos válidos para ajustar'}), 400
        
        ajustes_realizados = []
        
        for conteo in conteos_a_ajustar:
            if abs(conteo.diferencia) > 0.01:  # Solo ajustar si hay diferencia significativa
                # Actualizar el stock del artículo
                article = conteo.article
                stock_anterior = article.stock
                article.stock = conteo.cantidad_fisica
                
                # Registrar en historial de conteo físico
                observation = f"Ajuste por conteo físico - Sesión: {session_id}"
                if conteo.observaciones:
                    observation += f" - Obs: {conteo.observaciones}"
                
                log_physical_count(
                    article_id=article.id,
                    user_id=session['user_id'],
                    old_stock=stock_anterior,
                    new_stock=article.stock,
                    observation=observation
                )
                
                # Marcar el conteo como ajustado
                conteo.estado = 'ajustado'
                conteo.fecha_ajuste = datetime.utcnow()
                
                ajustes_realizados.append({
                    'article_id': article.id,
                    'article_title': article.title,
                    'stock_anterior': stock_anterior,
                    'stock_nuevo': article.stock,
                    'diferencia': conteo.diferencia,
                    'unit_type': getattr(article, 'unit_type', 'unidades')
                })
        
        db.session.commit()
        
        return jsonify({
            'message': f'Ajustes aplicados exitosamente a {len(ajustes_realizados)} productos',
            'ajustes': ajustes_realizados
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error al aplicar ajustes: {str(e)}")
        return jsonify({'error': 'Error al aplicar ajustes de inventario'}), 500

@app.route('/physical-inventory/sessions', methods=['GET'])
@login_required
def get_inventory_sessions():
    """Obtiene el historial de sesiones de conteo físico"""
    try:
        # Obtener sesiones únicas con información resumida
        sessions_data = db.session.query(
            PhysicalInventory.conteo_session_id,
            func.min(PhysicalInventory.fecha_conteo).label('fecha_conteo'),
            func.min(PhysicalInventory.usuario_conteo).label('usuario_conteo'),
            func.count(PhysicalInventory.id).label('total_productos'),
            func.sum(case((PhysicalInventory.estado == 'contado', 1), else_=0)).label('contados'),
            func.sum(case((PhysicalInventory.estado == 'ajustado', 1), else_=0)).label('ajustados'),
            func.sum(case((func.abs(PhysicalInventory.diferencia) > 0.01, 1), else_=0)).label('con_diferencias')
        ).group_by(
            PhysicalInventory.conteo_session_id
        ).order_by(func.min(PhysicalInventory.fecha_conteo).desc()).all()
        
        sessions = []
        for session in sessions_data:
            sessions.append({
                'session_id': session.conteo_session_id,
                'fecha_conteo': session.fecha_conteo.strftime('%Y-%m-%d %H:%M:%S'),
                'usuario_conteo': session.usuario_conteo,
                'total_productos': session.total_productos,
                'contados': session.contados,
                'ajustados': session.ajustados,
                'con_diferencias': session.con_diferencias,
                'estado': 'Completado' if session.contados == session.total_productos else 'En progreso'
            })
        
        return jsonify({'sessions': sessions})
        
    except Exception as e:
        print(f"Error al obtener sesiones: {str(e)}")
        return jsonify({'error': 'Error al obtener historial de sesiones'}), 500

# =====================
# DESCUENTOS Y PROMOCIONES
# =====================

@app.route('/discounts', methods=['GET'])
@login_required
def get_discounts():
    """Obtiene todos los descuentos disponibles"""
    try:
        discounts = Discount.query.order_by(Discount.created_at.desc()).all()
        return jsonify({
            'discounts': [discount.to_dict() for discount in discounts]
        })
    except Exception as e:
        print(f"Error al obtener descuentos: {str(e)}")
        return jsonify({'error': 'Error al obtener descuentos'}), 500

@app.route('/discounts/active', methods=['GET'])
@login_required
def get_active_discounts():
    """Obtiene solo los descuentos activos"""
    try:
        discounts = Discount.query.filter(Discount.activo == True).all()
        active_discounts = [d for d in discounts if d.is_active()]
        return jsonify({
            'discounts': [discount.to_dict() for discount in active_discounts]
        })
    except Exception as e:
        print(f"Error al obtener descuentos activos: {str(e)}")
        return jsonify({'error': 'Error al obtener descuentos activos'}), 500

@app.route('/promotions', methods=['GET'])
@login_required
def get_promotions():
    """Obtiene todas las promociones"""
    try:
        promotions = Promotion.query.order_by(Promotion.prioridad.desc(), Promotion.created_at.desc()).all()
        return jsonify({
            'promotions': [promotion.to_dict() for promotion in promotions]
        })
    except Exception as e:
        print(f"Error al obtener promociones: {str(e)}")
        return jsonify({'error': 'Error al obtener promociones'}), 500

@app.route('/promotions/active', methods=['GET'])
@login_required
def get_active_promotions():
    """Obtiene solo las promociones activas en este momento"""
    try:
        promotions = Promotion.query.filter(Promotion.activo == True).order_by(Promotion.prioridad.desc()).all()
        active_promotions = [p for p in promotions if p.is_active()]
        return jsonify({
            'promotions': [promotion.to_dict() for promotion in active_promotions]
        })
    except Exception as e:
        print(f"Error al obtener promociones activas: {str(e)}")
        return jsonify({'error': 'Error al obtener promociones activas'}), 500

@app.route('/promotions', methods=['POST'])
@login_required
def create_promotion():
    """Crea una nueva promoción"""
    try:
        data = request.get_json()
        
        # Validar campos requeridos
        required_fields = ['nombre', 'tipo', 'descuento_tipo', 'descuento_valor']
        for field in required_fields:
            if field not in data:
                return jsonify({'error': f'Campo requerido: {field}'}), 400
        
        # Crear nueva promoción
        promotion = Promotion(
            nombre=data['nombre'],
            descripcion=data.get('descripcion', ''),
            tipo=data['tipo'],
            condiciones=data.get('condiciones', '{}'),
            descuento_tipo=data['descuento_tipo'],
            descuento_valor=data['descuento_valor'],
            activo=data.get('activo', True),
            fecha_inicio=datetime.strptime(data['fecha_inicio'], '%Y-%m-%d') if data.get('fecha_inicio') else None,
            fecha_fin=datetime.strptime(data['fecha_fin'], '%Y-%m-%d') if data.get('fecha_fin') else None,
            dias_semana=data.get('dias_semana'),
            hora_inicio=datetime.strptime(data['hora_inicio'], '%H:%M').time() if data.get('hora_inicio') else None,
            hora_fin=datetime.strptime(data['hora_fin'], '%H:%M').time() if data.get('hora_fin') else None,
            usos_maximos_dia=data.get('usos_maximos_dia'),
            prioridad=data.get('prioridad', 1)
        )
        
        db.session.add(promotion)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Promoción creada exitosamente',
            'promotion': promotion.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error al crear promoción: {str(e)}")
        return jsonify({'error': 'Error al crear promoción'}), 500

@app.route('/promotions/<int:promotion_id>', methods=['PUT'])
@login_required
def update_promotion(promotion_id):
    """Actualiza una promoción existente"""
    try:
        promotion = Promotion.query.get(promotion_id)
        if not promotion:
            return jsonify({'error': 'Promoción no encontrada'}), 404
        
        data = request.get_json()
        
        # Actualizar campos
        if 'nombre' in data:
            promotion.nombre = data['nombre']
        if 'descripcion' in data:
            promotion.descripcion = data['descripcion']
        if 'tipo' in data:
            promotion.tipo = data['tipo']
        if 'condiciones' in data:
            promotion.condiciones = data['condiciones']
        if 'descuento_tipo' in data:
            promotion.descuento_tipo = data['descuento_tipo']
        if 'descuento_valor' in data:
            promotion.descuento_valor = data['descuento_valor']
        if 'activo' in data:
            promotion.activo = data['activo']
        if 'fecha_inicio' in data:
            promotion.fecha_inicio = datetime.strptime(data['fecha_inicio'], '%Y-%m-%d') if data['fecha_inicio'] else None
        if 'fecha_fin' in data:
            promotion.fecha_fin = datetime.strptime(data['fecha_fin'], '%Y-%m-%d') if data['fecha_fin'] else None
        if 'dias_semana' in data:
            promotion.dias_semana = data['dias_semana']
        if 'hora_inicio' in data:
            promotion.hora_inicio = datetime.strptime(data['hora_inicio'], '%H:%M').time() if data['hora_inicio'] else None
        if 'hora_fin' in data:
            promotion.hora_fin = datetime.strptime(data['hora_fin'], '%H:%M').time() if data['hora_fin'] else None
        if 'usos_maximos_dia' in data:
            promotion.usos_maximos_dia = data['usos_maximos_dia']
        if 'prioridad' in data:
            promotion.prioridad = data['prioridad']
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Promoción actualizada exitosamente',
            'promotion': promotion.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error al actualizar promoción: {str(e)}")
        return jsonify({'error': 'Error al actualizar promoción'}), 500

@app.route('/promotions/<int:promotion_id>', methods=['DELETE'])
@login_required
def delete_promotion(promotion_id):
    """Elimina una promoción"""
    try:
        promotion = Promotion.query.get(promotion_id)
        if not promotion:
            return jsonify({'error': 'Promoción no encontrada'}), 404
        
        db.session.delete(promotion)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Promoción eliminada exitosamente'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error al eliminar promoción: {str(e)}")
        return jsonify({'error': 'Error al eliminar promoción'}), 500

@app.route('/promotions/<int:promotion_id>/toggle', methods=['POST'])
@login_required
def toggle_promotion(promotion_id):
    """Activa o desactiva una promoción"""
    try:
        promotion = Promotion.query.get(promotion_id)
        if not promotion:
            return jsonify({'error': 'Promoción no encontrada'}), 404
        
        promotion.activo = not promotion.activo
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Promoción {"activada" if promotion.activo else "desactivada"} exitosamente',
            'promotion': promotion.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error al cambiar estado de promoción: {str(e)}")
        return jsonify({'error': 'Error al cambiar estado de promoción'}), 500

@app.route('/products/simple', methods=['GET'])
@login_required
def get_products_simple():
    """Obtiene lista simple de productos para uso en promociones"""
    try:
        products = Article.query.filter(Article.activo == True).all()
        return jsonify({
            'products': [{
                'id': product.id,
                'title': product.title,
                'precio': product.precio,
                'stock': product.stock,
                'unit_type': getattr(product, 'unit_type', 'unidades')
            } for product in products]
        })
    except Exception as e:
        print(f"Error al obtener productos: {str(e)}")
        return jsonify({'error': 'Error al obtener productos'}), 500

@app.route('/articles/most-sold', methods=['GET'])
def get_most_sold_products():
    """Obtiene los productos más vendidos basándose en las ventas"""
    try:
        print("=== Iniciando consulta de productos más vendidos ===")
        
        # Obtener top 3 productos más vendidos por cantidad total
        most_sold = db.session.query(
            SaleItem.article_id,
            func.sum(SaleItem.quantity).label('total_vendido'),
            func.count(SaleItem.id).label('veces_vendido'),
            Article.title,
            Article.precio,
            Article.content,
            Article.stock,
            Article.image_url,
            Article.unit_type,
            Article.margen_ganancia
        ).join(
            Article, SaleItem.article_id == Article.id
        ).filter(
            Article.activo == True
        ).group_by(
            SaleItem.article_id
        ).order_by(
            func.sum(SaleItem.quantity).desc()
        ).limit(3).all()
        
        print(f"Productos encontrados: {len(most_sold)}")
        
        products = []
        for item in most_sold:
            product_data = {
                'id': item.article_id,
                'title': item.title,
                'content': item.content or '',
                'precio': item.precio,
                'stock': item.stock,
                'image_url': item.image_url,
                'unit_type': getattr(item, 'unit_type', 'unidades'),
                'margen_ganancia': getattr(item, 'margen_ganancia', None),
                'total_vendido': float(item.total_vendido),
                'veces_vendido': item.veces_vendido,
                'activo': True
            }
            products.append(product_data)
            print(f"- {item.title}: {item.total_vendido} vendidos")
        
        print("=== Consulta completada exitosamente ===")
        return jsonify({
            'products': products
        })
        
    except Exception as e:
        print(f"Error al obtener productos más vendidos: {str(e)}")
        return jsonify({'error': 'Error al obtener productos más vendidos'}), 500

@app.route('/cart/apply-discount', methods=['POST'])
@login_required
def apply_manual_discount():
    """Aplica un descuento manual al carrito"""
    try:
        data = request.get_json()
        cart_items = data.get('cart_items', [])
        discount_type = data.get('discount_type')  # 'porcentaje' o 'cantidad_fija'
        discount_value = data.get('discount_value', 0)
        description = data.get('description', 'Descuento manual')
        
        if not cart_items:
            return jsonify({'error': 'Carrito vacío'}), 400
            
        if discount_type not in ['porcentaje', 'cantidad_fija']:
            return jsonify({'error': 'Tipo de descuento inválido'}), 400
            
        if discount_value <= 0:
            return jsonify({'error': 'Valor de descuento debe ser mayor a 0'}), 400
        
        # Calcular total del carrito
        subtotal = sum(float(item['precio']) * float(item['quantity']) for item in cart_items)
        
        # Calcular descuento
        if discount_type == 'porcentaje':
            if discount_value > 100:
                return jsonify({'error': 'Porcentaje no puede ser mayor a 100%'}), 400
            discount_amount = subtotal * (discount_value / 100)
        else:  # cantidad_fija
            if discount_value > subtotal:
                return jsonify({'error': 'Descuento no puede ser mayor al total'}), 400
            discount_amount = discount_value
        
        # Redondear a 2 decimales
        discount_amount = round(discount_amount, 2)
        new_total = round(subtotal - discount_amount, 2)
        
        return jsonify({
            'success': True,
            'discount': {
                'type': discount_type,
                'value': discount_value,
                'amount': discount_amount,
                'description': description,
                'percentage_applied': round((discount_amount / subtotal) * 100, 2) if subtotal > 0 else 0
            },
            'totals': {
                'subtotal': round(subtotal, 2),
                'discount_amount': discount_amount,
                'total': new_total
            }
        })
        
    except Exception as e:
        print(f"Error al aplicar descuento manual: {str(e)}")
        return jsonify({'error': 'Error al aplicar descuento'}), 500

@app.route('/cart/check-promotions', methods=['POST'])
@login_required
def check_promotions():
    """Verifica qué promociones se pueden aplicar al carrito actual"""
    try:
        data = request.get_json()
        cart_items = data.get('cart_items', [])
        
        if not cart_items:
            return jsonify({'applicable_promotions': []})
        
        # Obtener promociones activas
        promotions = Promotion.query.filter(Promotion.activo == True).order_by(Promotion.prioridad.desc()).all()
        active_promotions = [p for p in promotions if p.is_active()]
        
        applicable_promotions = []
        
        for promotion in active_promotions:
            is_applicable = check_promotion_conditions(promotion, cart_items)
            if is_applicable:
                discount_amount = calculate_promotion_discount(promotion, cart_items)
                affected_products = get_affected_products(promotion, cart_items)
                
                applicable_promotions.append({
                    'promotion': promotion.to_dict(),
                    'estimated_discount': discount_amount,
                    'affected_products': affected_products
                })
        
        return jsonify({
            'applicable_promotions': applicable_promotions
        })
        
    except Exception as e:
        print(f"Error al verificar promociones: {str(e)}")
        return jsonify({'error': 'Error al verificar promociones'}), 500

def check_promotion_conditions(promotion, cart_items):
    """Verifica si una promoción se puede aplicar a los items del carrito"""
    try:
        condiciones = promotion.get_condiciones()
        
        if promotion.tipo == 'descuento_general':
            # Verificar si hay mínimo de compra
            minimo_compra = condiciones.get('minimo_compra', 0)
            if minimo_compra > 0:
                total_carrito = sum(float(item['precio']) * float(item['quantity']) for item in cart_items)
                return total_carrito >= minimo_compra
            return True
            
        elif promotion.tipo == 'combo':
            # Verificar si todos los productos requeridos están en el carrito
            productos_requeridos = condiciones.get('productos', [])
            for producto_req in productos_requeridos:
                product_id = producto_req.get('id')
                cantidad_req = producto_req.get('cantidad', 1)
                
                # Buscar en el carrito
                found = False
                for item in cart_items:
                    if item['id'] == product_id and item['quantity'] >= cantidad_req:
                        found = True
                        break
                
                if not found:
                    return False
            
            return True
            
        elif promotion.tipo == 'descuento_cantidad':
            # Verificar cantidad mínima de productos específicos
            product_id = condiciones.get('product_id')
            cantidad_minima = condiciones.get('cantidad_minima', 1)
            
            if product_id:
                for item in cart_items:
                    if item['id'] == product_id and item['quantity'] >= cantidad_minima:
                        return True
                return False
            else:
                # Si no hay product_id específico, verificar cantidad total
                total_cantidad = sum(item['quantity'] for item in cart_items)
                return total_cantidad >= cantidad_minima
        
        return False
        
    except Exception as e:
        print(f"Error verificando condiciones: {str(e)}")
        return False

def calculate_promotion_discount(promotion, cart_items):
    """Calcula el descuento que aplicaría una promoción sobre productos específicos"""
    try:
        condiciones = promotion.get_condiciones()
        applicable_total = 0
        
        print(f"Calculando descuento para promoción: {promotion.nombre}")
        print(f"Tipo: {promotion.tipo}, Descuento: {promotion.descuento_valor}{'' if promotion.descuento_tipo == 'cantidad_fija' else '%'}")
        
        if promotion.tipo == 'descuento_general':
            # Para descuento general, aplicar a todo el carrito
            applicable_total = sum(float(item['precio']) * float(item['quantity']) for item in cart_items)
            print(f"Descuento general - Total aplicable: ${applicable_total}")
            
        elif promotion.tipo == 'combo':
            # Para combo, aplicar solo a los productos del combo
            productos_combo = condiciones.get('productos', [])
            print(f"Productos del combo requeridos: {productos_combo}")
            
            for producto_req in productos_combo:
                product_id = producto_req.get('id')
                cantidad_req = producto_req.get('cantidad', 1)
                
                # Buscar el producto en el carrito y calcular su costo
                for item in cart_items:
                    if item['id'] == product_id:
                        # Aplicar descuento solo a la cantidad requerida para el combo
                        cantidad_descuento = min(item['quantity'], cantidad_req)
                        costo_producto = float(item['precio']) * cantidad_descuento
                        applicable_total += costo_producto
                        print(f"Producto {item['title']}: ${item['precio']} x {cantidad_descuento} = ${costo_producto}")
                        break
            
            print(f"Total aplicable para combo: ${applicable_total}")
                        
        elif promotion.tipo == 'descuento_cantidad':
            product_id = condiciones.get('product_id')
            cantidad_minima = condiciones.get('cantidad_minima', 1)
            
            if product_id:
                # Aplicar solo al producto específico
                for item in cart_items:
                    if item['id'] == product_id and item['quantity'] >= cantidad_minima:
                        applicable_total = float(item['precio']) * item['quantity']
                        print(f"Descuento por cantidad - Producto específico: ${applicable_total}")
                        break
            else:
                # Si no hay product_id específico, aplicar a todos los productos
                # pero solo si se cumple la cantidad mínima total
                total_cantidad = sum(item['quantity'] for item in cart_items)
                if total_cantidad >= cantidad_minima:
                    applicable_total = sum(float(item['precio']) * float(item['quantity']) for item in cart_items)
                    print(f"Descuento por cantidad - Todos los productos: ${applicable_total}")
        
        # Calcular descuento según el tipo
        descuento_final = 0
        if promotion.descuento_tipo == 'porcentaje':
            descuento_final = applicable_total * (promotion.descuento_valor / 100)
            print(f"Descuento porcentual: ${applicable_total} x {promotion.descuento_valor}% = ${descuento_final}")
            
        elif promotion.descuento_tipo == 'cantidad_fija':
            # El descuento fijo no puede ser mayor al total aplicable
            descuento_final = min(promotion.descuento_valor, applicable_total)
            print(f"Descuento fijo: min(${promotion.descuento_valor}, ${applicable_total}) = ${descuento_final}")
        
        resultado = round(descuento_final, 2)
        print(f"Descuento final calculado: ${resultado}")
        return resultado
        
    except Exception as e:
        print(f"Error calculando descuento de promoción: {str(e)}")
        return 0

def get_affected_products(promotion, cart_items):
    """Obtiene la lista de productos afectados por una promoción"""
    try:
        condiciones = promotion.get_condiciones()
        affected_products = []
        
        if promotion.tipo == 'descuento_general':
            # Todos los productos del carrito están afectados
            affected_products = [{'id': item['id'], 'title': item['title']} for item in cart_items]
            
        elif promotion.tipo == 'combo':
            # Solo los productos del combo están afectados
            productos_combo = condiciones.get('productos', [])
            for producto_req in productos_combo:
                product_id = producto_req.get('id')
                
                # Buscar el producto en el carrito
                for item in cart_items:
                    if item['id'] == product_id:
                        affected_products.append({
                            'id': item['id'], 
                            'title': item['title'],
                            'cantidad_descuento': min(item['quantity'], producto_req.get('cantidad', 1))
                        })
                        break
                        
        elif promotion.tipo == 'descuento_cantidad':
            product_id = condiciones.get('product_id')
            
            if product_id:
                # Solo el producto específico está afectado
                for item in cart_items:
                    if item['id'] == product_id:
                        affected_products.append({'id': item['id'], 'title': item['title']})
                        break
            else:
                # Todos los productos están afectados
                affected_products = [{'id': item['id'], 'title': item['title']} for item in cart_items]
        
        return affected_products
        
    except Exception as e:
        print(f"Error obteniendo productos afectados: {str(e)}")
        return []

# =====================
# VENTAS
# =====================

@app.route('/sales', methods=['POST'])
@login_required
def create_sale():
    try:
        data = request.get_json()
        
        # Imprimir datos recibidos para debug
        print(f"Datos recibidos: {data}")
        
        cart_items = data.get('cart_items', [])
        metodo_pago = data.get('metodo_pago', 'efectivo')
        suspended_sale_id = data.get('suspended_sale_id')
        
        # Datos de descuentos y promociones
        applied_discount = data.get('applied_discount')  # Descuento manual aplicado
        applied_promotions = data.get('applied_promotions', [])  # Promociones aplicadas
        
        # Nota de la venta
        nota = data.get('nota')  # Nota para la venta
        
        if not cart_items:
            return jsonify({'error': 'No hay items en el carrito'}), 400
        
        # Verificar estructura de cart_items
        for item in cart_items:
            required_fields = ['id', 'title', 'precio', 'quantity']
            missing_fields = [field for field in required_fields if field not in item]
            if missing_fields:
                return jsonify({'error': f'Faltan campos en item: {missing_fields}'}), 400
        
        # Obtener turno activo
        turno_activo = Turno.query.filter_by(user_id=session['user_id'], activo=True).first()
        if not turno_activo:
            return jsonify({'error': 'No hay turno activo'}), 400
        
        # Validar stock según tipo de producto
        for item in cart_items:
            article = Article.query.get(item['id'])
            if not article:
                return jsonify({'error': f'Producto {item["title"]} no encontrado'}), 404
            
            unit_type = getattr(article, 'unit_type', 'unidades')
            requested_quantity = item['quantity']
            
            if unit_type == 'peso':
                # Para productos por peso, validar que hay suficiente stock en kg
                if article.stock < requested_quantity:
                    return jsonify({
                        'error': f'Stock insuficiente para {article.title}. Stock disponible: {article.stock}kg, solicitado: {requested_quantity}kg'
                    }), 400
            else:
                # Para productos por unidades, validar como antes
                if article.stock < requested_quantity:
                    return jsonify({
                        'error': f'Stock insuficiente para {article.title}. Stock disponible: {article.stock} unidades'
                    }), 400
        
        # Calcular subtotal
        subtotal = sum(item['precio'] * item['quantity'] for item in cart_items)
        
        # Calcular descuentos totales
        total_discount = 0
        
        # Aplicar descuento manual si existe
        if applied_discount:
            if applied_discount['type'] == 'porcentaje':
                manual_discount = subtotal * (applied_discount['value'] / 100)
            else:  # cantidad_fija
                manual_discount = applied_discount['value']
            total_discount += manual_discount
        
        # Aplicar promociones si existen
        for promotion_data in applied_promotions:
            promotion_discount = promotion_data.get('discount_amount', 0)
            total_discount += promotion_discount
        
        # Calcular total final
        total = round(subtotal - total_discount, 2)
        
        # Asegurar que el total no sea negativo
        if total < 0:
            total = 0
        
        # Generar número de ticket
        ticket_number = f"T{turno_activo.id}-{uuid.uuid4().hex[:6].upper()}"
        
        # Crear venta
        nueva_venta = Sale(
            ticket_number=ticket_number,
            total=total,
            metodo_pago=metodo_pago,
            turno_id=turno_activo.id,
            user_id=session['user_id'],
            nota=nota  # Incluir la nota
        )
        db.session.add(nueva_venta)
        db.session.flush()  # Para obtener el ID
        
        # Registrar descuentos aplicados
        if applied_discount:
            sale_discount = SaleDiscount(
                sale_id=nueva_venta.id,
                tipo_descuento='manual',
                descripcion=applied_discount.get('description', 'Descuento manual'),
                monto_descuento=applied_discount.get('amount', manual_discount),
                porcentaje_aplicado=applied_discount.get('value') if applied_discount['type'] == 'porcentaje' else None,
                aplicado_por=session.get('username', 'Sistema')
            )
            db.session.add(sale_discount)
        
        # Registrar promociones aplicadas
        for promotion_data in applied_promotions:
            promotion = promotion_data.get('promotion', {})
            sale_discount = SaleDiscount(
                sale_id=nueva_venta.id,
                tipo_descuento='promocion',
                promotion_id=promotion.get('id'),
                descripcion=promotion.get('nombre', 'Promoción aplicada'),
                monto_descuento=promotion_data.get('discount_amount', 0),
                porcentaje_aplicado=None,
                aplicado_por=session.get('username', 'Sistema')
            )
            db.session.add(sale_discount)
        
        # Crear items de venta y actualizar stock
        for item in cart_items:
            # Crear item de venta
            sale_item = SaleItem(
                sale_id=nueva_venta.id,
                article_id=item['id'],
                article_title=item['title'],  # Guardar título
                quantity=item['quantity'],
                unit_price=item['precio'],
                subtotal=item['precio'] * item['quantity']
            )
            db.session.add(sale_item)
            
            # Actualizar stock
            article = Article.query.get(item['id'])
            if article:
                article.stock -= item['quantity']
        
        # Si es una venta retomada, eliminar la venta suspendida
        if suspended_sale_id:
            suspended_sale = SuspendedSale.query.get(suspended_sale_id)
            if suspended_sale:
                db.session.delete(suspended_sale)
        
        # Actualizar estadísticas del turno
        turno_activo.cantidad_ventas = (turno_activo.cantidad_ventas or 0) + 1
        turno_activo.total_ventas = (turno_activo.total_ventas or 0) + total
        
        if metodo_pago == 'efectivo':
            turno_activo.total_efectivo = (turno_activo.total_efectivo or 0) + total
        else:
            turno_activo.total_tarjeta = (turno_activo.total_tarjeta or 0) + total
        
        db.session.commit()
        
        # Actualizar la categoría de productos frecuentes después de cada venta
        try:
            update_frequent_products_category()
        except Exception as e:
            print(f"Error al actualizar productos frecuentes: {e}")
            # No fallar la venta si hay error en la actualización de frecuentes
        
        return jsonify({
            'success': True,
            'message': 'Venta procesada exitosamente',
            'ticket_number': ticket_number,
            'sale_id': nueva_venta.id,
            'subtotal': round(subtotal, 2),
            'total_discount': round(total_discount, 2),
            'total': total
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error detallado en venta: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Error al procesar la venta: {str(e)}'}), 500

@app.route('/sales', methods=['GET'])
@login_required
def get_sales():
    try:
        turno_activo = Turno.query.filter_by(user_id=session['user_id'], activo=True).first()
        
        if not turno_activo:
            return jsonify({
                'sales': [],
                'total_pages': 0,
                'current_page': 1,
                'total_sales': 0,
                'message': 'No hay turno activo'
            })
        
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 10, type=int)
        
        sales = Sale.query.filter_by(turno_id=turno_activo.id).order_by(Sale.fecha_venta.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        sales_data = []
        for sale in sales.items:
            sale_data = {
                'id': sale.id,
                'ticket_number': sale.ticket_number,
                'total': sale.total,
                'metodo_pago': sale.metodo_pago,
                'fecha_venta': sale.fecha_venta.strftime('%Y-%m-%d %H:%M:%S'),
                'user': sale.user.username if sale.user else 'Unknown',
                'items_count': len(sale.items),
                'nota': sale.nota,  # Incluir la nota de la venta
                'items': [
                    {
                        'article_title': item.article_title,
                        'quantity': item.quantity,
                        'unit_price': item.unit_price,
                        'subtotal': item.subtotal
                    }
                    for item in sale.items
                ]
            }
            sales_data.append(sale_data)
        
        return jsonify({
            'sales': sales_data,
            'total_pages': sales.pages,
            'current_page': sales.page,
            'total_sales': sales.total,
            'turno_info': {
                'id': turno_activo.id,
                'fecha_inicio': turno_activo.fecha_inicio.strftime('%Y-%m-%d %H:%M:%S'),
                'total_efectivo': turno_activo.total_efectivo or 0,
                'total_tarjeta': turno_activo.total_tarjeta or 0,
                'total_ventas': turno_activo.total_ventas or 0,
                'cantidad_ventas': turno_activo.cantidad_ventas or 0
            }
        })
        
    except Exception as e:
        return jsonify({'error': 'Error al obtener ventas'}), 500

@app.route('/sales/<int:sale_id>/note', methods=['PUT'])
@permission_required('can_add_notes')
def add_sale_note(sale_id):
    """Agregar o editar nota en una venta"""
    try:
        data = request.get_json()
        nota = data.get('nota', '').strip()
        
        # Buscar la venta
        sale = Sale.query.get(sale_id)
        if not sale:
            return jsonify({'error': 'Venta no encontrada'}), 404
        
        # Verificar que la venta pertenezca al turno activo del usuario
        turno_activo = Turno.query.filter_by(user_id=session['user_id'], activo=True).first()
        if not turno_activo or sale.turno_id != turno_activo.id:
            return jsonify({'error': 'No puedes editar ventas de otros turnos'}), 403
        
        # Actualizar la nota
        sale.nota = nota if nota else None
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Nota guardada exitosamente',
            'nota': sale.nota
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al guardar la nota: {str(e)}'}), 500

# =====================
# VENTAS SUSPENDIDAS
# =====================

@app.route('/sales/suspend', methods=['POST'])
@login_required
def suspend_sale():
    try:
        data = request.get_json()
        cart_items = data.get('cart_items', [])
        nota = data.get('nota', '')
        
        if not cart_items:
            return jsonify({'error': 'No hay items para suspender'}), 400
        
        # Obtener turno activo
        turno_activo = Turno.query.filter_by(user_id=session['user_id'], activo=True).first()
        if not turno_activo:
            return jsonify({'error': 'No hay turno activo'}), 400
        
        # Calcular total
        total = sum(item['precio'] * item['quantity'] for item in cart_items)
        
        # Generar número de ticket único
        ticket_number = f"SUSP-{uuid.uuid4().hex[:8].upper()}"
        
        # Crear venta suspendida
        suspended_sale = SuspendedSale(
            turno_id=turno_activo.id,
            user_id=session['user_id'],
            ticket_number=ticket_number,
            total=total,
            items_data=json.dumps(cart_items),
            nota=nota
        )
        
        db.session.add(suspended_sale)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Venta suspendida exitosamente',
            'ticket_number': ticket_number,
            'suspended_sale_id': suspended_sale.id
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error suspendiendo venta: {str(e)}")
        return jsonify({'error': 'Error al suspender venta'}), 500

@app.route('/sales/suspended', methods=['GET'])
@login_required
def get_suspended_sales():
    try:
        # Obtener turno activo
        turno_activo = Turno.query.filter_by(user_id=session['user_id'], activo=True).first()
        if not turno_activo:
            return jsonify([])
        
        # Obtener ventas suspendidas
        suspended_sales = SuspendedSale.query.filter_by(turno_id=turno_activo.id)\
                                           .order_by(SuspendedSale.fecha_suspension.desc())\
                                           .all()
        
        sales_data = []
        for sale in suspended_sales:
            items = json.loads(sale.items_data)
            user = User.query.get(sale.user_id)
            
            sales_data.append({
                'id': sale.id,
                'ticket_number': sale.ticket_number,
                'total': sale.total,
                'items_count': len(items),
                'items': items,
                'fecha_suspension': sale.fecha_suspension.isoformat(),
                'nota': sale.nota,
                'usuario': user.username if user else 'Desconocido'
            })
        
        return jsonify(sales_data)
        
    except Exception as e:
        print(f"Error obteniendo ventas suspendidas: {str(e)}")
        return jsonify({'error': 'Error al obtener ventas suspendidas'}), 500

@app.route('/sales/suspended/<int:suspended_id>/resume', methods=['POST'])
@login_required
def resume_suspended_sale(suspended_id):
    try:
        # Obtener venta suspendida
        suspended_sale = SuspendedSale.query.get_or_404(suspended_id)
        
        # Verificar que pertenece al turno activo del usuario
        turno_activo = Turno.query.filter_by(user_id=session['user_id'], activo=True).first()
        if not turno_activo or suspended_sale.turno_id != turno_activo.id:
            return jsonify({'error': 'Venta suspendida no válida'}), 400
        
        # Obtener items de la venta suspendida
        items = json.loads(suspended_sale.items_data)
        
        return jsonify({
            'success': True,
            'message': 'Venta lista para retomar',
            'cart_items': items,
            'ticket_number': suspended_sale.ticket_number,
            'suspended_sale_id': suspended_sale.id
        })
        
    except Exception as e:
        print(f"Error retomando venta: {str(e)}")
        return jsonify({'error': 'Error al retomar venta'}), 500

@app.route('/sales/suspended/<int:suspended_id>', methods=['DELETE'])
@login_required
def delete_suspended_sale(suspended_id):
    try:
        # Obtener y eliminar
        suspended_sale = SuspendedSale.query.get_or_404(suspended_id)
        
        # Verificar que pertenece al usuario actual
        turno_activo = Turno.query.filter_by(user_id=session['user_id'], activo=True).first()
        if not turno_activo or suspended_sale.turno_id != turno_activo.id:
            return jsonify({'error': 'No autorizado'}), 403
        
        db.session.delete(suspended_sale)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Venta suspendida eliminada exitosamente'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error eliminando venta suspendida: {str(e)}")
        return jsonify({'error': 'Error al eliminar venta suspendida'}), 500

# =====================
# DEVOLUCIONES
# =====================

@app.route('/returns', methods=['POST'])
@permission_required('can_process_returns')
def create_return():
    try:
        data = request.get_json()
        article_id = data.get('article_id')
        quantity = data.get('quantity')
        motivo = data.get('motivo', '').strip()
        
        if not article_id or not quantity or not motivo:
            return jsonify({'error': 'Faltan datos requeridos'}), 400
        
        if quantity <= 0:
            return jsonify({'error': 'La cantidad debe ser mayor a 0'}), 400
        
        # Obtener turno activo
        turno_activo = Turno.query.filter_by(user_id=session['user_id'], activo=True).first()
        if not turno_activo:
            return jsonify({'error': 'No hay turno activo'}), 400
        
        # Obtener artículo
        article = Article.query.get(article_id)
        if not article:
            return jsonify({'error': 'Artículo no encontrado'}), 404
        
        # Calcular total de la devolución
        total = article.precio * quantity
        
        # Generar número de ticket único para devolución
        ticket_number = f"DEV-{uuid.uuid4().hex[:8].upper()}"
        
        # Crear devolución
        nueva_devolucion = Devolucion(
            turno_id=turno_activo.id,
            user_id=session['user_id'],
            ticket_number=ticket_number,
            article_id=article.id,
            article_title=article.title,
            quantity=quantity,
            unit_price=article.precio,
            total=total,
            motivo=motivo
        )
        db.session.add(nueva_devolucion)
        
        # Actualizar stock del artículo (devolver al inventario)
        article.stock += quantity
        
        # Actualizar estadísticas del turno
        turno_activo.total_devoluciones = (turno_activo.total_devoluciones or 0) + total
        turno_activo.cantidad_devoluciones = (turno_activo.cantidad_devoluciones or 0) + 1
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Devolución procesada exitosamente',
            'ticket_number': ticket_number,
            'devolucion_id': nueva_devolucion.id,
            'total': total,
            'nuevo_stock': article.stock
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error en devolución: {str(e)}")
        return jsonify({'error': f'Error al procesar la devolución: {str(e)}'}), 500

@app.route('/returns', methods=['GET'])
@permission_required('can_process_returns')
def get_returns():
    try:
        turno_activo = Turno.query.filter_by(user_id=session['user_id'], activo=True).first()
        if not turno_activo:
            return jsonify([])
        
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 10, type=int)
        
        devoluciones = Devolucion.query.filter_by(turno_id=turno_activo.id)\
                                     .order_by(Devolucion.fecha_devolucion.desc())\
                                     .paginate(page=page, per_page=per_page, error_out=False)
        
        devoluciones_data = []
        for devolucion in devoluciones.items:
            devolucion_data = {
                'id': devolucion.id,
                'ticket_number': devolucion.ticket_number,
                'article_title': devolucion.article_title,
                'quantity': devolucion.quantity,
                'unit_price': devolucion.unit_price,
                'total': devolucion.total,
                'motivo': devolucion.motivo,
                'fecha_devolucion': devolucion.fecha_devolucion.strftime('%Y-%m-%d %H:%M:%S'),
                'usuario': devolucion.user.username if devolucion.user else 'Desconocido'
            }
            devoluciones_data.append(devolucion_data)
        
        return jsonify({
            'returns': devoluciones_data,
            'total_pages': devoluciones.pages,
            'current_page': devoluciones.page,
            'total_returns': devoluciones.total
        })
        
    except Exception as e:
        print(f"Error obteniendo devoluciones: {str(e)}")
        return jsonify({'error': 'Error al obtener devoluciones'}), 500

@app.route('/returns/<int:turno_id>', methods=['GET'])
@permission_required('can_process_returns')
def get_returns_by_turno(turno_id):
    try:
        # Verificar que el turno existe
        turno = Turno.query.get(turno_id)
        if not turno:
            return jsonify({'error': 'Turno no encontrado'}), 404
        
        # Obtener devoluciones del turno
        devoluciones = Devolucion.query.filter_by(turno_id=turno_id)\
                                     .order_by(Devolucion.fecha_devolucion.desc())\
                                     .all()
        
        devoluciones_data = []
        for devolucion in devoluciones:
            devolucion_data = {
                'id': devolucion.id,
                'ticket_number': devolucion.ticket_number,
                'article_title': devolucion.article_title,
                'quantity': devolucion.quantity,
                'unit_price': devolucion.unit_price,
                'total': devolucion.total,
                'motivo': devolucion.motivo,
                'fecha_devolucion': devolucion.fecha_devolucion.isoformat(),
                'usuario': devolucion.user.username if devolucion.user else 'Desconocido'
            }
            devoluciones_data.append(devolucion_data)
        
        return jsonify(devoluciones_data)
        
    except Exception as e:
        print(f"Error obteniendo devoluciones del turno: {str(e)}")
        return jsonify({'error': 'Error al obtener devoluciones del turno'}), 500

# =====================
# TURNOS
# =====================

@app.route('/close-turno', methods=['POST'])
@login_required
def close_turno():
    try:
        turno_activo = Turno.query.filter_by(user_id=session['user_id'], activo=True).first()
        
        if not turno_activo:
            return jsonify({'error': 'No hay turno activo'}), 400
        
        # Cerrar turno usando la zona horaria de Chile
        chile_tz = pytz.timezone('America/Santiago')
        turno_activo.fecha_cierre = datetime.now(chile_tz)
        turno_activo.activo = False
        
        db.session.commit()
        
        # Preparar resumen del turno
        resumen = {
            'turno_id': turno_activo.id,
            'fecha_inicio': turno_activo.fecha_inicio.strftime('%Y-%m-%d %H:%M:%S'),
            'fecha_cierre': turno_activo.fecha_cierre.strftime('%Y-%m-%d %H:%M:%S'),
            'total_ventas': float(turno_activo.total_ventas or 0),
            'total_efectivo': float(turno_activo.total_efectivo or 0),
            'total_tarjeta': float(turno_activo.total_tarjeta or 0),
            'cantidad_ventas': turno_activo.cantidad_ventas or 0,
            'usuario': turno_activo.user.username if turno_activo.user else 'Desconocido'
        }
        
        return jsonify({
            'message': 'Turno cerrado exitosamente',
            'resumen': resumen
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error al cerrar turno: {str(e)}")
        return jsonify({'error': 'Error interno al cerrar turno'}), 500

@app.route('/turno-actual', methods=['GET'])
@login_required
def get_turno_actual():
    try:
        turno_activo = Turno.query.filter_by(user_id=session['user_id'], activo=True).first()
        
        if not turno_activo:
            return jsonify({'turno_activo': False})
        
        return jsonify({
            'turno_activo': True,
            'turno_id': turno_activo.id,
            'fecha_inicio': turno_activo.fecha_inicio.strftime('%Y-%m-%d %H:%M:%S'),
            'total_ventas': turno_activo.total_ventas or 0,
            'total_efectivo': turno_activo.total_efectivo or 0,
            'total_tarjeta': turno_activo.total_tarjeta or 0,
            'cantidad_ventas': turno_activo.cantidad_ventas or 0,
            'total_devoluciones': turno_activo.total_devoluciones or 0,  # ← AGREGAR
            'cantidad_devoluciones': turno_activo.cantidad_devoluciones or 0  # ← AGREGAR
        })
        
    except Exception as e:
        return jsonify({'error': 'Error al obtener turno'}), 500

@app.route('/historial-turnos', methods=['GET'])
@permission_required('can_view_shift_history')
def get_historial_turnos():
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 10, type=int)
        
        turnos = Turno.query.filter_by(activo=False).order_by(Turno.fecha_cierre.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        turnos_data = []
        for turno in turnos.items:
            turno_data = {
                'id': turno.id,
                'usuario': turno.user.username,
                'fecha_inicio': turno.fecha_inicio.strftime('%Y-%m-%d %H:%M:%S'),
                'fecha_cierre': turno.fecha_cierre.strftime('%Y-%m-%d %H:%M:%S') if turno.fecha_cierre else None,
                'total_ventas': turno.total_ventas or 0,
                'total_efectivo': turno.total_efectivo or 0,
                'total_tarjeta': turno.total_tarjeta or 0,
                'cantidad_ventas': turno.cantidad_ventas or 0
            }
            turnos_data.append(turno_data)
        
        return jsonify({
            'turnos': turnos_data,
            'total_pages': turnos.pages,
            'current_page': turnos.page,
            'total_turnos': turnos.total
        })
        
    except Exception as e:
        return jsonify({'error': 'Error al obtener historial de turnos'}), 500

@app.route('/turnos-graficos', methods=['GET'])
@permission_required('can_view_shift_history')
def get_turnos_graficos():
    try:
        # Obtener parámetros de fecha
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')

        print(f"Fechas recibidas: inicio={fecha_inicio}, fin={fecha_fin}")

        # Query base para ventas y cantidades
        query = db.session.query(
            func.strftime('%Y-%m-%d', Sale.fecha_venta).label('fecha'),
            func.sum(Sale.total).label('total_ventas'),
            func.count(Sale.id).label('cantidad_total')
        ).filter(Sale.turno_id.isnot(None))

        # Aplicar filtros de fecha si se proporcionan
        if fecha_inicio:
            try:
                fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d')
                query = query.filter(Sale.fecha_venta >= fecha_inicio)
                print(f"Filtro fecha inicio aplicado: {fecha_inicio}")
            except ValueError as e:
                print(f"Error al parsear fecha_inicio: {e}")
                return jsonify({'error': 'Formato de fecha inicio inválido'}), 400
        
        if fecha_fin:
            try:
                fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d')
                fecha_fin = fecha_fin + timedelta(days=1)  # Incluir todo el día final
                query = query.filter(Sale.fecha_venta < fecha_fin)
                print(f"Filtro fecha fin aplicado: {fecha_fin}")
            except ValueError as e:
                print(f"Error al parsear fecha_fin: {e}")
                return jsonify({'error': 'Formato de fecha fin inválido'}), 400

        # Agrupar por fecha y ordenar
        query = query.group_by(func.strftime('%Y-%m-%d', Sale.fecha_venta))\
                     .order_by(func.strftime('%Y-%m-%d', Sale.fecha_venta))
        
        print("Ejecutando consulta...")
        resultados = query.all()
        print(f"Resultados obtenidos: {len(resultados)}")
        
        # Debug: imprimir los tipos de datos y valores
        for idx, r in enumerate(resultados):
            print(f"Resultado {idx}:")
            print(f"  fecha: {type(r[0])} = {r[0]}")
            print(f"  ventas: {type(r[1])} = {r[1]}")
            print(f"  cantidad: {type(r[2])} = {r[2]}")

        # Preparar datos para los gráficos
        labels = []
        ventas_data = []
        cantidad_data = []
        
        for r in resultados:
            # La fecha viene como string desde la consulta SQL
            fecha = r[0]
            if isinstance(fecha, datetime):
                labels.append(fecha.strftime('%Y-%m-%d'))
            else:
                labels.append(str(fecha))
            
            # Convertir valores numéricos de manera segura
            try:
                ventas_data.append(float(r[1]) if r[1] is not None else 0.0)
            except (TypeError, ValueError):
                ventas_data.append(0.0)
                
            try:
                cantidad_data.append(int(r[2]) if r[2] is not None else 0)
            except (TypeError, ValueError):
                cantidad_data.append(0)

        # Retornar datos en formato JSON para los gráficos
        return jsonify({
            'ventas': {
                'labels': labels,
                'data': ventas_data
            },
            'cantidad': {
                'labels': labels,
                'data': cantidad_data
            }
        })

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error obteniendo datos para gráficos: {str(e)}")
        print(f"Detalles del error:\n{error_details}")
        return jsonify({'error': str(e)}), 500

@app.route('/turnos-graficos/vista', methods=['GET'])
@permission_required('can_view_shift_history')
def ver_graficos_turnos():
    return render_template('graficos_turnos.html')

@app.route('/turnos-historial', methods=['GET'])
@permission_required('can_view_shift_history')
def get_turnos_historial():
    try:
        # Obtener parámetros de fecha
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        
        # Configurar la zona horaria de Chile/Santiago
        chile_tz = pytz.timezone('America/Santiago')
        
        # Query base
        query = db.session.query(
            Turno,
            User,
            func.coalesce(func.sum(Sale.total), 0).label('total_ventas'),
            func.count(Sale.id).label('num_ventas'),
            func.coalesce(func.sum(case((Devolucion.id != None, Devolucion.total), else_=0)), 0).label('total_devoluciones'),
            func.count(Devolucion.id).label('num_devoluciones')
        ).join(User, Turno.user_id == User.id)\
         .outerjoin(Sale, Sale.turno_id == Turno.id)\
         .outerjoin(Devolucion, Devolucion.turno_id == Turno.id)
        
        # Aplicar filtros de fecha si se proporcionan
        if fecha_inicio:
            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            query = query.filter(Turno.fecha_inicio >= fecha_inicio)

        if fecha_fin:
            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d')
            fecha_fin = fecha_fin + timedelta(days=1)  # Incluir todo el día final
            query = query.filter(Turno.fecha_inicio < fecha_fin)
        
        # Agrupar y ordenar resultados
        results = query.group_by(Turno.id, User.id)\
                      .order_by(Turno.fecha_inicio.desc())\
                      .all()        # Formatear resultados
        turnos_list = []
        for turno, user, total_ventas, num_ventas, total_devoluciones, num_devoluciones in results:
            fecha_apertura_str = None
            fecha_cierre_str = None
            
            if turno.fecha_inicio:
                fecha_apertura_str = turno.fecha_inicio.astimezone(chile_tz).strftime('%Y-%m-%d %H:%M:%S')
            
            if turno.fecha_cierre:
                fecha_cierre_str = turno.fecha_cierre.astimezone(chile_tz).strftime('%Y-%m-%d %H:%M:%S')
            
            turno_dict = {
                'turno_id': turno.id,
                'usuario': user.username,
                'fecha_inicio': fecha_apertura_str,
                'fecha_cierre': fecha_cierre_str,
                'activo': turno.activo,
                'total_ventas': float(total_ventas),
                'num_ventas': num_ventas or 0,
                'total_devoluciones': float(total_devoluciones),
                'num_devoluciones': num_devoluciones or 0,
                'total_efectivo': float(turno.total_efectivo),
                'total_tarjeta': float(turno.total_tarjeta),
                'cantidad_ventas': int(turno.cantidad_ventas),
                'cantidad_devoluciones': int(turno.cantidad_devoluciones)
            }
            turnos_list.append(turno_dict)
        
        return jsonify(turnos_list)
        
    except Exception as e:
        print(f"Error en get_turnos_historial: {str(e)}")
        return jsonify({'error': str(e)}), 500
        
        if fecha_inicio and fecha_fin:
            query = query.filter(
                Turno.fecha_inicio >= fecha_inicio,
                Turno.fecha_cierre <= fecha_fin
            )
        
        turnos = query.group_by(Turno.id)\
                      .order_by(Turno.fecha_inicio.desc())\
                      .all()
        
        return jsonify([{
            'turno_id': turno.turno_id,
            'fecha_inicio': turno.fecha_inicio.isoformat(),
            'fecha_cierre': turno.fecha_cierre.isoformat() if turno.fecha_cierre else None,
            'usuario': turno.usuario,
            'total_ventas': float(turno.total_ventas),
            'total_efectivo': float(turno.total_efectivo),
            'total_tarjeta': float(turno.total_tarjeta),
            'cantidad_ventas': turno.cantidad_ventas,
            'total_devoluciones': float(turno.total_devoluciones),
            'cantidad_devoluciones': turno.cantidad_devoluciones
        } for turno in turnos])
        
    except Exception as e:
        print(f"Error en turnos-historial: {str(e)}")
        return jsonify({'error': 'Error al obtener historial de turnos'}), 500

@app.route('/turnos-historial/excel', methods=['GET'])
@permission_required('can_view_shift_history')
def export_turnos_excel():
    temp_path = None
    try:
        # Limpiar archivos temporales antiguos
        def remove_old_files():
            try:
                for filename in os.listdir(UPLOAD_FOLDER):
                    if filename.startswith('turnos_historial_') and filename.endswith('.xlsx'):
                        file_path = os.path.join(UPLOAD_FOLDER, filename)
                        try:
                            if os.path.isfile(file_path):
                                os.unlink(file_path)
                        except PermissionError:
                            # Si el archivo está en uso, lo intentamos ignorar
                            continue
                        except Exception as e:
                            # Log otros errores pero continuamos
                            print(f'Error al eliminar archivo {filename}: {e}')
            except Exception as e:
                print(f'Error al limpiar archivos temporales: {e}')
        
        # Intentar limpiar archivos antiguos, pero no bloquear si falla
        remove_old_files()

        # Obtener parámetros de fecha
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')

        # Crear un nuevo libro de trabajo
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historial de Turnos"

        # Configurar estilos
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')
        
        # Configurar zona horaria
        chile_tz = pytz.timezone('America/Santiago')

        # Definir anchos de columna
        column_widths = {
            1: 8,   # ID
            2: 15,  # Usuario
            3: 20,  # Fecha Apertura
            4: 20,  # Fecha Cierre
            5: 15,  # Total Ventas
            6: 18,  # Ventas en Efectivo
            7: 18,  # Ventas con Tarjeta
            8: 12,  # Num. Ventas
            9: 18,  # Total Devoluciones
            10: 15, # Num. Devoluciones
            11: 10  # Estado
        }
        
        # Escribir encabezados y configurar columnas
        headers = ['ID', 'Usuario', 'Fecha Apertura', 'Fecha Cierre', 'Total Ventas', 'Ventas en Efectivo', 
                  'Ventas con Tarjeta', 'Num. Ventas', 'Total Devoluciones', 'Num. Devoluciones', 'Estado']
                  
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.alignment = header_alignment
            
            # Establecer ancho de columna
            ws.column_dimensions[get_column_letter(col)].width = column_widths[col]
            
        # Estilo para las celdas con números
        number_alignment = Alignment(horizontal='right')

        # Query base
        query = db.session.query(
            Turno,
            User,
            func.coalesce(func.sum(Sale.total), 0).label('total_ventas'),
            func.count(Sale.id).label('num_ventas'),
            func.coalesce(func.sum(case((Devolucion.id != None, Devolucion.total), else_=0)), 0).label('total_devoluciones'),
            func.count(Devolucion.id).label('num_devoluciones')
        ).join(User, Turno.user_id == User.id)\
         .outerjoin(Sale, Sale.turno_id == Turno.id)\
         .outerjoin(Devolucion, Devolucion.turno_id == Turno.id)

        # Aplicar filtros de fecha si se proporcionan
        if fecha_inicio:
            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            query = query.filter(Turno.fecha_inicio >= fecha_inicio)
        
        if fecha_fin:
            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d')
            fecha_fin = fecha_fin + timedelta(days=1)
            query = query.filter(Turno.fecha_inicio < fecha_fin)

        # Agrupar y ordenar
        query = query.group_by(Turno.id, User.id)\
                    .order_by(Turno.fecha_inicio.desc())

        # Escribir datos
        for row, (turno, user, total_ventas, num_ventas, total_devoluciones, num_devoluciones) in enumerate(query.all(), 2):
            # ID y Usuario
            ws.cell(row=row, column=1, value=turno.id)
            ws.cell(row=row, column=2, value=user.username)
            
            # Fechas con formato
            cell_fecha_inicio = ws.cell(row=row, column=3)
            fecha_inicio_local = turno.fecha_inicio.astimezone(chile_tz).replace(tzinfo=None)
            cell_fecha_inicio.value = fecha_inicio_local
            cell_fecha_inicio.number_format = 'yyyy-mm-dd hh:mm:ss'
            
            cell_fecha_cierre = ws.cell(row=row, column=4)
            if turno.fecha_cierre:
                fecha_cierre_local = turno.fecha_cierre.astimezone(chile_tz).replace(tzinfo=None)
                cell_fecha_cierre.value = fecha_cierre_local
                cell_fecha_cierre.number_format = 'yyyy-mm-dd hh:mm:ss'
            else:
                cell_fecha_cierre.value = "Abierto"
            
            # Valores monetarios con formato
            cell_total_ventas = ws.cell(row=row, column=5, value=float(total_ventas if total_ventas is not None else 0.0))
            cell_total_ventas.number_format = '#,##0.00'
            cell_total_ventas.alignment = number_alignment
            
            cell_total_efectivo = ws.cell(row=row, column=6, value=float(turno.total_efectivo or 0.0))
            cell_total_efectivo.number_format = '#,##0.00'
            cell_total_efectivo.alignment = number_alignment
            
            cell_total_tarjeta = ws.cell(row=row, column=7, value=float(turno.total_tarjeta or 0.0))
            cell_total_tarjeta.number_format = '#,##0.00'
            cell_total_tarjeta.alignment = number_alignment
            
            # Cantidades
            cell_num_ventas = ws.cell(row=row, column=8, value=int(num_ventas or 0))
            cell_num_ventas.number_format = '#,##0'
            cell_num_ventas.alignment = number_alignment
            
            cell_total_devol = ws.cell(row=row, column=9, value=float(total_devoluciones or 0.0))
            cell_total_devol.number_format = '#,##0.00'
            cell_total_devol.alignment = number_alignment
            
            cell_num_devol = ws.cell(row=row, column=10, value=int(num_devoluciones or 0))
            cell_num_devol.number_format = '#,##0'
            cell_num_devol.alignment = number_alignment
            
            # Estado
            ws.cell(row=row, column=11, value="Activo" if turno.activo else "Cerrado")
            ws.cell(row=row, column=8, value=int(num_ventas or 0)).number_format = '#,##0'
            ws.cell(row=row, column=9, value=float(total_devoluciones or 0.0)).number_format = '#,##0.00'
            ws.cell(row=row, column=10, value=int(num_devoluciones or 0)).number_format = '#,##0'
            
            # Estado
            ws.cell(row=row, column=11, value="Activo" if turno.activo else "Cerrado")

        # Ajustar ancho de columnas y aplicar formatos
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

            # Aplicar formato de moneda a columnas específicas
            if column[0].column in [5, 6, 7, 9]:  # Columnas con valores monetarios
                for cell in column[1:]:  # Excluir el encabezado
                    cell.number_format = '#,##0.00'

        try:
            # Guardar archivo temporalmente con nombre único
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
            temp_filename = f"turnos_historial_{timestamp}.xlsx"
            temp_path = os.path.join(UPLOAD_FOLDER, temp_filename)
            
            # Intentar guardar el archivo
            max_retries = 3
            retry_delay = 0.5  # segundos
            
            for attempt in range(max_retries):
                try:
                    wb.save(temp_path)
                    # Si el guardado fue exitoso, salimos del bucle
                    break
                except PermissionError as e:
                    if attempt < max_retries - 1:
                        time.sleep(retry_delay)
                        # Intentar con un nuevo nombre si falla
                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
                        temp_filename = f"turnos_historial_{timestamp}.xlsx"
                        temp_path = os.path.join(UPLOAD_FOLDER, temp_filename)
                    else:
                        raise e
            if not os.path.exists(temp_path):
                raise Exception("El archivo no se pudo crear correctamente")

            # Enviar archivo
            return send_file(
                temp_path,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='historial_turnos.xlsx'
            )
        except Exception as e:
            print(f"Error al guardar o enviar el archivo Excel: {str(e)}")
            return jsonify({'error': 'Error al exportar a Excel'}), 500
        finally:
            # Intentar eliminar el archivo temporal después de enviarlo
            try:
                if os.path.exists(temp_path):
                    os.unlink(temp_path)
            except Exception as e:
                print(f"Error al eliminar archivo temporal: {str(e)}")
    
    except Exception as e:
        print(f"Error en export_turnos_excel: {str(e)}")
        return jsonify({'error': str(e)}), 500
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/turnos-historial/pdf', methods=['GET'])
@permission_required('can_view_shift_history')
def export_turnos_pdf():
    try:
        # Obtener parámetros de fecha
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')

        # Configurar zona horaria
        chile_tz = pytz.timezone('America/Santiago')

        # Crear documento PDF
        temp_filename = f"turnos_historial_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        temp_path = os.path.join(app.config['UPLOAD_FOLDER'], temp_filename)
        doc = SimpleDocTemplate(temp_path, pagesize=letter)

        # Crear elementos del PDF
        elements = []
        
        # Estilo para el título
        styles = getSampleStyleSheet()
        title_style = styles['Heading1']
        title = Paragraph("Historial de Turnos", title_style)
        elements.append(title)

        # Query base
        query = db.session.query(
            Turno,
            User,
            func.coalesce(func.sum(Sale.total), 0).label('total_ventas'),
            func.count(Sale.id).label('num_ventas'),
            func.coalesce(func.sum(case((Devolucion.id != None, Devolucion.total), else_=0)), 0).label('total_devoluciones'),
            func.count(Devolucion.id).label('num_devoluciones')
        ).join(User, Turno.user_id == User.id)\
         .outerjoin(Sale, Sale.turno_id == Turno.id)\
         .outerjoin(Devolucion, Devolucion.turno_id == Turno.id)

        # Aplicar filtros de fecha si se proporcionan
        if fecha_inicio:
            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            query = query.filter(Turno.fecha_inicio >= fecha_inicio)
        
        if fecha_fin:
            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d')
            fecha_fin = fecha_fin + timedelta(days=1)
            query = query.filter(Turno.fecha_inicio < fecha_fin)

        # Agrupar y ordenar
        query = query.group_by(Turno.id, User.id)\
                    .order_by(Turno.fecha_inicio.desc())

        # Crear tabla de datos
        data = [['ID', 'Usuario', 'Fecha Apertura', 'Fecha Cierre', 'Total Ventas', 
                'Num. Ventas', 'Total Dev.', 'Num. Dev.', 'Estado']]

        for turno, user, total_ventas, num_ventas, total_devoluciones, num_devoluciones in query.all():
            row = [
                str(turno.id),
                user.username,
                turno.fecha_inicio.astimezone(chile_tz).strftime('%Y-%m-%d %H:%M:%S'),
                turno.fecha_cierre.astimezone(chile_tz).strftime('%Y-%m-%d %H:%M:%S') if turno.fecha_cierre else "Abierto",
                f"${float(total_ventas):,.0f}" if total_ventas else "$0",
                str(num_ventas),
                f"${float(total_devoluciones):,.0f}" if total_devoluciones else "$0",
                str(num_devoluciones),
                "Activo" if turno.activo else "Cerrado"
            ]
            data.append(row)

        # Crear tabla y establecer estilo
        table = Table(data)
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BOX', (0, 0), (-1, -1), 2, colors.black),
            ('GRID', (0, 0), (-1, 0), 2, colors.black)
        ])
        table.setStyle(style)
        
        # Añadir tabla al documento
        elements.append(table)
        
        # Generar PDF
        doc.build(elements)

        # Enviar archivo
        return send_file(
            temp_path,
            mimetype='application/pdf',
            as_attachment=True,
            download_name='historial_turnos.pdf'
        )

    except Exception as e:
        print(f"Error en export_turnos_pdf: {str(e)}")
        return jsonify({'error': str(e)}), 500
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
        
        query = db.session.query(
            Turno.id.label('turno_id'),
            Turno.fecha_inicio,
            Turno.fecha_cierre,
            User.username.label('usuario'),
            func.coalesce(func.sum(Sale.total), 0).label('total_ventas'),
            func.coalesce(func.sum(case((Sale.metodo_pago == 'efectivo', Sale.total), else_=0)), 0).label('total_efectivo'),
            func.coalesce(func.sum(case((Sale.metodo_pago == 'tarjeta', Sale.total), else_=0)), 0).label('total_tarjeta'),
            func.coalesce(func.count(Sale.id), 0).label('cantidad_ventas'),
            # Campos de devoluciones
            func.coalesce(Turno.total_devoluciones, 0).label('total_devoluciones'),
            func.coalesce(Turno.cantidad_devoluciones, 0).label('cantidad_devoluciones')
        ).join(User, Turno.user_id == User.id)\
         .outerjoin(Sale, Turno.id == Sale.turno_id)\
         .filter(Turno.activo == False)
        
        if fecha_inicio and fecha_fin:
            query = query.filter(
                func.date(Turno.fecha_inicio) >= fecha_inicio,
                func.date(Turno.fecha_inicio) <= fecha_fin
            )
        
        turnos = query.group_by(Turno.id)\
                      .order_by(Turno.fecha_inicio.desc())\
                      .all()
        
        return jsonify([{
            'turno_id': turno.turno_id,
            'fecha_inicio': turno.fecha_inicio.isoformat(),
            'fecha_cierre': turno.fecha_cierre.isoformat() if turno.fecha_cierre else None,
            'usuario': turno.usuario,
            'total_ventas': float(turno.total_ventas),
            'total_efectivo': float(turno.total_efectivo),
            'total_tarjeta': float(turno.total_tarjeta),
            'cantidad_ventas': turno.cantidad_ventas,
            'total_devoluciones': float(turno.total_devoluciones),  
            'cantidad_devoluciones': turno.cantidad_devoluciones  
        } for turno in turnos])
        
    except Exception as e:
        print(f"Error en turnos-historial: {str(e)}")
        return jsonify({'error': 'Error al obtener historial de turnos'}), 500

@app.route('/turnos/<int:turno_id>/ventas', methods=['GET'])
@permission_required('can_view_shift_history')
def get_ventas_turno(turno_id):
    try:
        # Verificar que el turno existe
        turno = Turno.query.get(turno_id)
        if not turno:
            return jsonify({'error': 'Turno no encontrado'}), 404
        
        # Obtener ventas del turno
        ventas = Sale.query.filter_by(turno_id=turno_id)\
                          .order_by(Sale.fecha_venta.desc())\
                          .all()
        
        result = []
        for venta in ventas:
            # Obtener items de la venta
            items = SaleItem.query.filter_by(sale_id=venta.id).all()
            
            venta_data = {
                'id': venta.id,
                'ticket_number': venta.ticket_number,
                'total': float(venta.total),
                'metodo_pago': venta.metodo_pago,
                'fecha_venta': venta.fecha_venta.isoformat(),
                'user': venta.user.username if venta.user else 'Desconocido',
                'items_count': len(items),
                'nota': venta.nota,  # Incluir la nota de la venta
                'items': []
            }
            
            # Agregar items de la venta
            for item in items:
                item_data = {
                    'article_title': item.article_title,
                    'quantity': item.quantity,
                    'unit_price': float(item.unit_price),
                    'subtotal': float(item.subtotal)
                }
                venta_data['items'].append(item_data)
            
            result.append(venta_data)
        
        return jsonify(result)
        
    except Exception as e:
        print(f"Error al obtener ventas del turno: {str(e)}")
        return jsonify({'error': 'Error al obtener ventas del turno'}), 500

# =====================
# HISTORIALES DE AUDITORÍA
# =====================

@app.route('/history/products', methods=['GET'])
@permission_required('can_view_audit_logs')
def get_product_history():
    """Obtener historial de cambios en productos"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        article_id = request.args.get('article_id', type=int)
        action = request.args.get('action')
        user_id = request.args.get('user_id', type=int)
        
        # Construir query base
        query = ProductHistory.query
        
        # Filtros opcionales
        if article_id:
            query = query.filter(ProductHistory.article_id == article_id)
        if action:
            query = query.filter(ProductHistory.action == action)
        if user_id:
            query = query.filter(ProductHistory.user_id == user_id)
        
        # Paginar y ordenar por fecha descendente
        history = query.order_by(ProductHistory.timestamp.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        result = []
        for entry in history.items:
            entry_data = {
                'id': entry.id,
                'article_id': entry.article_id,
                'article_title': entry.article.title if entry.article else 'Producto eliminado',
                'user_id': entry.user_id,
                'username': entry.user.username if entry.user else 'Usuario desconocido',
                'action': entry.action,
                'description': entry.description,
                'old_values': json.loads(entry.old_values) if entry.old_values else None,
                'new_values': json.loads(entry.new_values) if entry.new_values else None,
                'timestamp': entry.timestamp.strftime('%Y-%m-%d %H:%M:%S')
            }
            result.append(entry_data)
        
        return jsonify({
            'history': result,
            'total_pages': history.pages,
            'current_page': history.page,
            'total_entries': history.total,
            'has_next': history.has_next,
            'has_prev': history.has_prev
        })
        
    except Exception as e:
        print(f"Error obteniendo historial de productos: {str(e)}")
        return jsonify({'error': 'Error al obtener historial de productos'}), 500

@app.route('/history/physical-counts', methods=['GET'])
@permission_required('can_perform_physical_count')
def get_physical_count_history():
    """Obtener historial de conteos físicos"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        article_id = request.args.get('article_id', type=int)
        user_id = request.args.get('user_id', type=int)
        
        # Construir query base
        query = PhysicalCountHistory.query
        
        # Filtros opcionales
        if article_id:
            query = query.filter(PhysicalCountHistory.article_id == article_id)
        if user_id:
            query = query.filter(PhysicalCountHistory.user_id == user_id)
        
        # Paginar y ordenar por fecha descendente
        history = query.order_by(PhysicalCountHistory.timestamp.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        result = []
        for entry in history.items:
            entry_data = {
                'id': entry.id,
                'article_id': entry.article_id,
                'article_title': entry.article.title if entry.article else 'Producto eliminado',
                'user_id': entry.user_id,
                'username': entry.user.username if entry.user else 'Usuario desconocido',
                'old_stock': entry.old_stock,
                'new_stock': entry.new_stock,
                'difference': entry.difference,
                'observation': entry.observation,
                'timestamp': entry.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                'unit_type': getattr(entry.article, 'unit_type', 'unidades') if entry.article else 'unidades'
            }
            result.append(entry_data)
        
        return jsonify({
            'history': result,
            'total_pages': history.pages,
            'current_page': history.page,
            'total_entries': history.total,
            'has_next': history.has_next,
            'has_prev': history.has_prev
        })
        
    except Exception as e:
        print(f"Error obteniendo historial de conteos físicos: {str(e)}")
        return jsonify({'error': 'Error al obtener historial de conteos físicos'}), 500

@app.route('/history/products/<int:article_id>', methods=['GET'])
@permission_required('can_view_audit_logs')
def get_product_specific_history(article_id):
    """Obtener historial específico de un producto"""
    try:
        # Verificar que el producto existe
        article = Article.query.get(article_id)
        if not article:
            return jsonify({'error': 'Producto no encontrado'}), 404
        
        # Obtener historial de cambios del producto
        product_changes = ProductHistory.query.filter_by(article_id=article_id).order_by(
            ProductHistory.timestamp.desc()
        ).all()
        
        # Obtener historial de conteos físicos del producto
        physical_counts = PhysicalCountHistory.query.filter_by(article_id=article_id).order_by(
            PhysicalCountHistory.timestamp.desc()
        ).all()
        
        changes_data = []
        for change in product_changes:
            changes_data.append({
                'type': 'product_change',
                'id': change.id,
                'user': change.user.username if change.user else 'Usuario desconocido',
                'action': change.action,
                'description': change.description,
                'old_values': json.loads(change.old_values) if change.old_values else None,
                'new_values': json.loads(change.new_values) if change.new_values else None,
                'timestamp': change.timestamp.strftime('%Y-%m-%d %H:%M:%S')
            })
        
        counts_data = []
        for count in physical_counts:
            counts_data.append({
                'type': 'physical_count',
                'id': count.id,
                'user': count.user.username if count.user else 'Usuario desconocido',
                'old_stock': count.old_stock,
                'new_stock': count.new_stock,
                'difference': count.difference,
                'observation': count.observation,
                'timestamp': count.timestamp.strftime('%Y-%m-%d %H:%M:%S')
            })
        
        return jsonify({
            'article': {
                'id': article.id,
                'title': article.title,
                'current_stock': article.stock,
                'unit_type': getattr(article, 'unit_type', 'unidades')
            },
            'changes': changes_data,
            'physical_counts': counts_data
        })
        
    except Exception as e:
        print(f"Error obteniendo historial específico: {str(e)}")
        return jsonify({'error': 'Error al obtener historial del producto'}), 500

# =====================
# GESTIÓN DE USUARIOS
# =====================

@app.route('/users', methods=['GET'])
@admin_required
def get_users():
    """Obtener lista de todos los usuarios (solo administradores)"""
    try:
        users = User.query.all()
        users_data = []
        
        for user in users:
            user_data = {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'is_admin': user.is_admin,
                'is_active': user.is_active,
                'created_at': user.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'permissions': user.get_permissions()
            }
            users_data.append(user_data)
        
        return jsonify({'users': users_data})
        
    except Exception as e:
        print(f"Error obteniendo usuarios: {str(e)}")
        return jsonify({'error': 'Error al obtener usuarios'}), 500

@app.route('/users', methods=['POST'])
@admin_required
def create_user():
    """Crear nuevo usuario (solo administradores)"""
    try:
        data = request.get_json()
        
        # Validar campos requeridos
        required_fields = ['username', 'email', 'password']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'Campo {field} es requerido'}), 400
        
        # Verificar si el usuario ya existe
        if User.query.filter_by(username=data['username']).first():
            return jsonify({'error': 'El nombre de usuario ya existe'}), 400
        
        if User.query.filter_by(email=data['email']).first():
            return jsonify({'error': 'El email ya está registrado'}), 400
        
        # Crear nuevo usuario
        new_user = User(
            username=data['username'],
            email=data['email'],
            is_admin=data.get('is_admin', False),
            is_active=data.get('is_active', True)
        )
        new_user.set_password(data['password'])
        
        # Asignar permisos
        permissions = data.get('permissions', {})
        new_user.update_permissions(permissions)
        
        db.session.add(new_user)
        db.session.commit()
        
        return jsonify({
            'message': 'Usuario creado exitosamente',
            'user': {
                'id': new_user.id,
                'username': new_user.username,
                'email': new_user.email,
                'is_admin': new_user.is_admin,
                'permissions': new_user.get_permissions()
            }
        }), 201
        
    except Exception as e:
        print(f"Error creando usuario: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Error al crear usuario'}), 500

@app.route('/users/<int:user_id>', methods=['PUT'])
@admin_required
def update_user(user_id):
    """Actualizar usuario existente (solo administradores)"""
    try:
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'Usuario no encontrado'}), 404
        
        data = request.get_json()
        
        # Actualizar campos básicos
        if 'username' in data:
            # Verificar que el nuevo username no esté tomado
            existing_user = User.query.filter_by(username=data['username']).first()
            if existing_user and existing_user.id != user_id:
                return jsonify({'error': 'El nombre de usuario ya existe'}), 400
            user.username = data['username']
        
        if 'email' in data:
            # Verificar que el nuevo email no esté tomado
            existing_user = User.query.filter_by(email=data['email']).first()
            if existing_user and existing_user.id != user_id:
                return jsonify({'error': 'El email ya está registrado'}), 400
            user.email = data['email']
        
        if 'password' in data and data['password']:
            user.set_password(data['password'])
        
        if 'is_admin' in data:
            user.is_admin = data['is_admin']
        
        if 'is_active' in data:
            user.is_active = data['is_active']
        
        # Actualizar permisos
        if 'permissions' in data:
            user.update_permissions(data['permissions'])
        
        db.session.commit()
        
        return jsonify({
            'message': 'Usuario actualizado exitosamente',
            'user': {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'is_admin': user.is_admin,
                'is_active': user.is_active,
                'permissions': user.get_permissions()
            }
        })
        
    except Exception as e:
        print(f"Error actualizando usuario: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Error al actualizar usuario'}), 500

@app.route('/users/<int:user_id>', methods=['DELETE'])
@admin_required
def delete_user(user_id):
    """Eliminar usuario (solo administradores)"""
    try:
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'Usuario no encontrado'}), 404
        
        # No permitir eliminar el último administrador
        if user.is_admin:
            admin_count = User.query.filter_by(is_admin=True).count()
            if admin_count <= 1:
                return jsonify({'error': 'No se puede eliminar el último administrador'}), 400
        
        # No permitir que un usuario se elimine a sí mismo
        current_user_id = session.get('user_id')
        if user_id == current_user_id:
            return jsonify({'error': 'No puedes eliminar tu propia cuenta'}), 400
        
        db.session.delete(user)
        db.session.commit()
        
        return jsonify({'message': 'Usuario eliminado exitosamente'})
        
    except Exception as e:
        print(f"Error eliminando usuario: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Error al eliminar usuario'}), 500

@app.route('/user/permissions', methods=['GET'])
@login_required
def get_user_permissions():
    """Obtener permisos del usuario actual"""
    try:
        user = User.query.get(session['user_id'])
        if not user:
            return jsonify({'error': 'Usuario no válido'}), 401
        
        return jsonify({
            'permissions': user.get_permissions(),
            'user_info': {
                'id': user.id,
                'username': user.username,
                'is_admin': user.is_admin
            }
        })
        
    except Exception as e:
        print(f"Error obteniendo permisos: {str(e)}")
        return jsonify({'error': 'Error al obtener permisos'}), 500

if __name__ == '__main__':
    app.run(host='localhost', port=5000, debug=True)
